from app import create_app
from app.blueprints.admin.shift_import import ShiftImportService
from io import BytesIO
import openpyxl

app = create_app()
with app.app_context():
    # Create a test Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Write headers
    ws['A1'] = 'EMP-CODE'
    ws['B1'] = 'HOSPITAL NAME'
    ws['C1'] = 'SHIFT'
    
    # Write sample data
    ws['A2'] = 'E-2512012'
    ws['B2'] = 'AIIMS Hospital (Gorakhpur)'
    ws['C2'] = '10:00 AM to 06:00 PM'
    
    ws['A3'] = 'E-2603025'
    ws['B3'] = 'Akurdi Hospital'
    ws['C3'] = '12:00 PM to 09:00 PM'
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create a mock FileStorage object
    class MockFileStorage:
        def __init__(self, stream):
            self.stream = stream
            self.filename = 'test.xlsx'
        
        def read(self):
            return self.stream.getvalue()
        
        def seek(self, pos):
            self.stream.seek(pos)
        
        def stream_method(self):
            return self.stream
    
    from werkzeug.datastructures import FileStorage
    file_obj = FileStorage(
        stream=output,
        filename='test.xlsx',
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Test the import
    service = ShiftImportService()
    result = service.import_from_file(file_obj, effective_date='2026-08-06', assigned_by_user_id=1)
    
    print("\n=== IMPORT RESULT ===")
    print(f"Success: {result.get('success')}")
    print(f"Message: {result.get('message')}")
    print(f"Hospitals Assigned: {result.get('hospitals_assigned')}")
    print(f"Shifts Assigned: {result.get('assigned')}")
    print(f"Not Found: {result.get('notfound')}")
    print(f"Errors: {result.get('errors')}")
    
    print("\n=== DETAILS ===")
    for detail in result.get('details', []):
        print(f"  - {detail['emp_code']}: {detail['status']} (hospital={detail['hospital_name']}, shift={detail['shift_name']})")
