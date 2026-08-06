from app import create_app
from app.blueprints.admin.shift_import import ShiftImportService
from io import BytesIO
import openpyxl

app = create_app()
with app.app_context():
    # Create test Excel with employees that EXIST
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Write headers
    ws['A1'] = 'EMP-CODE'
    ws['B1'] = 'HOSPITAL NAME'
    ws['C1'] = 'SHIFT'
    
    # Write test data with EXISTING employees and various hospital formats
    test_data = [
        ('E-2512012', 'AIIMS Hospital (Gorakhpur)', '8:00 AM to 5:00 PM'),
        ('E-2603025', 'Claim Team', '10:00 AM to 7:00 PM'),  # New hospital we just added
        ('E-2512012', 'DR. RN COOPER HOSPITAL', '12:00 PM to 9:00 AM'),  # Name mapping test
    ]
    
    for idx, (emp_code, hospital, shift) in enumerate(test_data, start=2):
        ws[f'A{idx}'] = emp_code
        ws[f'B{idx}'] = hospital
        ws[f'C{idx}'] = shift
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create a mock FileStorage object
    from werkzeug.datastructures import FileStorage
    file_obj = FileStorage(
        stream=output,
        filename='test_bulk.xlsx',
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Test the import
    service = ShiftImportService()
    result = service.import_from_file(file_obj, effective_date='2026-08-06', assigned_by_user_id=1)
    
    print("\n" + "="*60)
    print("IMPORT RESULT")
    print("="*60)
    print(f"Success: {result.get('success')}")
    print(f"Message: {result.get('message')}")
    print(f"Shifts Assigned: {result.get('assigned')}")
    print(f"Hospitals Assigned: {result.get('hospitals_assigned')}")
    print(f"Not Found: {result.get('notfound')}")
    print(f"Errors: {result.get('errors')}")
    
    print("\n" + "="*60)
    print("DETAILS")
    print("="*60)
    for detail in result.get('details', []):
        status_icon = "[OK]" if detail['status'] == 'assigned' else "[!]"
        print(f"{status_icon} {detail['emp_code']:12s} -> Hospital: {detail['hospital_name']:40s} | Shift: {detail['shift_name']}")

