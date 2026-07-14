"""
app/utils/export_utils.py
==========================
Data export utilities — CSV, Excel (XLSX), and PDF generation.

All export functions return (file_bytes, mimetype, filename) tuples
so that routes can stream them directly without temporary files.

Dependencies:
    csv       — stdlib, always available
    openpyxl  — for Excel exports
    reportlab — for PDF exports
"""

import csv
import io
from datetime import datetime, timezone
from typing import Any, Optional


def export_to_csv(
    headers: list[str],
    rows: list[list[Any]],
    filename: Optional[str] = None,
) -> tuple[bytes, str, str]:
    """
    Export tabular data to a CSV byte stream.

    Args:
        headers: Column header labels.
        rows: List of row value lists.
        filename: Desired download filename (without extension).

    Returns:
        Tuple of (csv_bytes, mimetype, filename_with_extension).
    """
    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_ALL)
    writer.writerow(headers)
    for row in rows:
        writer.writerow([str(v) if v is not None else "" for v in row])

    csv_bytes = output.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility
    fname = _build_filename(filename, "csv")
    return csv_bytes, "text/csv", fname


def export_to_excel(
    headers: list[str],
    rows: list[list[Any]],
    sheet_name: str = "Report",
    filename: Optional[str] = None,
) -> tuple[bytes, str, str]:
    """
    Export tabular data to an Excel XLSX byte stream.

    Args:
        headers: Column header labels.
        rows: List of row value lists.
        sheet_name: Name for the Excel worksheet.
        filename: Desired download filename (without extension).

    Returns:
        Tuple of (xlsx_bytes, mimetype, filename_with_extension).
    """
    try:
        from openpyxl import Workbook  # noqa: PLC0415
        from openpyxl.styles import Font, PatternFill, Alignment  # noqa: PLC0415
    except ImportError as exc:
        raise ImportError("openpyxl is required for Excel exports.") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Style header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A3C6E", end_color="1A3C6E", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Write data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-fit column widths (approximate)
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    xlsx_bytes = output.getvalue()
    fname = _build_filename(filename, "xlsx")
    return (
        xlsx_bytes,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        fname,
    )


def export_to_pdf(
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    filename: Optional[str] = None,
    page_size: str = "A4",
    landscape: bool = False,
) -> tuple[bytes, str, str]:
    """
    Export tabular data to a PDF byte stream using ReportLab.

    Args:
        title: Report title displayed at the top of the PDF.
        headers: Column header labels.
        rows: List of row value lists.
        filename: Desired download filename (without extension).
        page_size: 'A4' or 'letter'.
        landscape: If True, use landscape orientation.

    Returns:
        Tuple of (pdf_bytes, mimetype, filename_with_extension).
    """
    try:
        from reportlab.lib import colors  # noqa: PLC0415
        from reportlab.lib.pagesizes import A4, letter, landscape as rl_landscape  # noqa: PLC0415
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer  # noqa: PLC0415
        from reportlab.lib.styles import getSampleStyleSheet  # noqa: PLC0415
        from reportlab.lib.units import inch  # noqa: PLC0415
    except ImportError as exc:
        raise ImportError("reportlab is required for PDF exports.") from exc

    output = io.BytesIO()
    page = A4 if page_size.upper() == "A4" else letter
    if landscape:
        page = rl_landscape(page)

    doc = SimpleDocTemplate(output, pagesize=page,
                             leftMargin=0.75 * inch, rightMargin=0.75 * inch,
                             topMargin=inch, bottomMargin=0.75 * inch)
    styles = getSampleStyleSheet()
    story = []

    # Title
    story.append(Paragraph(title, styles["Title"]))
    story.append(Spacer(1, 0.25 * inch))

    # Generated timestamp
    generated = datetime.now(timezone.utc).strftime("%d %b %Y, %I:%M %p UTC")
    story.append(Paragraph(f"Generated: {generated}", styles["Normal"]))
    story.append(Spacer(1, 0.2 * inch))

    # Table data
    table_data = [headers] + [[str(v) if v is not None else "" for v in row] for row in rows]

    tbl = Table(table_data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A3C6E")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F6F9")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DEE2E6")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ]))

    story.append(tbl)
    doc.build(story)

    fname = _build_filename(filename, "pdf")
    return output.getvalue(), "application/pdf", fname


def stream_file_response(file_bytes: bytes, mimetype: str, filename: str):
    """
    Build a Flask streaming response for a file download.

    Args:
        file_bytes: Raw file content.
        mimetype: MIME type string.
        filename: Suggested download filename.

    Returns:
        Flask Response with Content-Disposition header.
    """
    from flask import Response  # noqa: PLC0415
    response = Response(
        file_bytes,
        mimetype=mimetype,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(file_bytes)),
        },
    )
    return response


def _build_filename(name: Optional[str], extension: str) -> str:
    """Build a sanitized filename with timestamp suffix."""
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    base = name or "export"
    # Remove unsafe characters
    base = "".join(c for c in base if c.isalnum() or c in ("_", "-")).strip("_-")
    return f"{base}_{timestamp}.{extension}"
