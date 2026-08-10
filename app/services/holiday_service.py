"""
app/services/holiday_service.py
=================================
Service for managing company holidays and Excel import.

Handles:
    - Reading holidays from Excel
    - Validating holiday data
    - Detecting and handling duplicates
    - Importing into database
"""

import logging
from datetime import date, datetime
from typing import Optional

import openpyxl
from werkzeug.datastructures import FileStorage

from app.extensions.database import db
from app.models.company_holiday import CompanyHoliday

logger = logging.getLogger(__name__)


class HolidayService:
    """Service for holiday management and Excel import."""

    ALLOWED_EXTENSIONS = {"xlsx"}

    def get_holidays_by_year(self, year: int) -> list[CompanyHoliday]:
        """
        Get all holidays for a specific year, sorted by date.

        Args:
            year: Year to fetch holidays for

        Returns:
            List of CompanyHoliday objects for the year
        """
        start_date = date(year, 1, 1)
        end_date = date(year, 12, 31)

        holidays = (
            CompanyHoliday.query
            .filter(
                CompanyHoliday.holiday_date >= start_date,
                CompanyHoliday.holiday_date <= end_date,
                CompanyHoliday.is_deleted == False,
            )
            .order_by(CompanyHoliday.holiday_date)
            .all()
        )
        return holidays

    def get_holiday_by_id(self, holiday_id: int) -> Optional[CompanyHoliday]:
        """
        Get a specific holiday by ID.

        Args:
            holiday_id: ID of the holiday

        Returns:
            CompanyHoliday object or None if not found
        """
        return (
            CompanyHoliday.query
            .filter_by(id=holiday_id, is_deleted=False)
            .first()
        )

    def get_available_years(self) -> list[int]:
        """
        Get list of years that have holidays.

        Returns:
            Sorted list of years with holidays
        """
        result = db.session.query(
            db.func.extract('year', CompanyHoliday.holiday_date).cast(db.Integer)
        ).filter(
            CompanyHoliday.is_deleted == False
        ).distinct().all()

        years = sorted([row[0] for row in result if row[0]])

        # Ensure current year and next year are always available
        current_year = datetime.now().year
        if current_year not in years:
            years.append(current_year)
        if current_year + 1 not in years:
            years.append(current_year + 1)

        return sorted(years)

    def import_from_excel(self, file: FileStorage) -> dict:
        """
        Import holidays from an Excel file.

        Expected columns (case-insensitive):
            - holiday_date or date: The date of the holiday (required)
            - holiday_name or name: Name of the holiday (required)
            - holiday_type or type: Type of holiday (optional)
            - description or desc: Description (optional)

        Args:
            file: FileStorage object from form

        Returns:
            {
                success: bool,
                message: str,
                added: int,
                skipped: int,
                errors: list of error messages,
            }
        """
        try:
            # Validate file
            if not file or file.filename == "":
                return {
                    "success": False,
                    "message": "No file selected.",
                    "added": 0,
                    "skipped": 0,
                    "errors": [],
                }

            if not self._allowed_file(file.filename):
                return {
                    "success": False,
                    "message": "File must be an Excel file (.xlsx).",
                    "added": 0,
                    "skipped": 0,
                    "errors": [],
                }

            # Read Excel
            workbook = openpyxl.load_workbook(file)
            worksheet = workbook.active

            if not worksheet:
                return {
                    "success": False,
                    "message": "Excel file is empty or invalid.",
                    "added": 0,
                    "skipped": 0,
                    "errors": [],
                }

            # Parse headers
            headers = self._parse_headers(worksheet)
            if not headers:
                return {
                    "success": False,
                    "message": "Could not find valid columns. Expected: holiday_date, holiday_name",
                    "added": 0,
                    "skipped": 0,
                    "errors": [],
                }

            # Validate required columns
            if "holiday_date" not in headers or "holiday_name" not in headers:
                return {
                    "success": False,
                    "message": "Missing required columns: holiday_date and holiday_name",
                    "added": 0,
                    "skipped": 0,
                    "errors": [],
                }

            # Process rows
            added = 0
            skipped = 0
            errors = []

            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=False), start=2):
                # Check if row is empty
                if all(cell.value is None for cell in row):
                    continue

                holiday_data = self._parse_row(row, headers, row_idx)

                if "error" in holiday_data:
                    errors.append(holiday_data["error"])
                    skipped += 1
                    continue

                # Check for duplicate
                existing = CompanyHoliday.query.filter_by(
                    holiday_date=holiday_data["holiday_date"],
                    holiday_name=holiday_data["holiday_name"],
                    is_deleted=False,
                ).first()

                if existing:
                    # Skip duplicate
                    skipped += 1
                    continue

                # Create new holiday
                try:
                    holiday = CompanyHoliday(
                        holiday_date=holiday_data["holiday_date"],
                        holiday_name=holiday_data["holiday_name"],
                        holiday_type=holiday_data.get("holiday_type"),
                        description=holiday_data.get("description"),
                    )
                    db.session.add(holiday)
                    added += 1
                except Exception as e:
                    logger.error(f"Error creating holiday from row {row_idx}: {e}")
                    errors.append(f"Row {row_idx}: Failed to create holiday")
                    skipped += 1

            # Commit all changes
            try:
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                logger.error(f"Error committing holidays: {e}")
                return {
                    "success": False,
                    "message": f"Database error: {str(e)[:100]}",
                    "added": 0,
                    "skipped": 0,
                    "errors": errors,
                }

            message = f"Successfully imported {added} holiday(ies)"
            if skipped > 0:
                message += f". {skipped} row(s) skipped (duplicates or invalid)."

            return {
                "success": True,
                "message": message,
                "added": added,
                "skipped": skipped,
                "errors": errors,
            }

        except Exception as e:
            logger.error(f"Error in import_from_excel: {e}")
            return {
                "success": False,
                "message": f"Error processing file: {str(e)[:100]}",
                "added": 0,
                "skipped": 0,
                "errors": [str(e)],
            }

    def _allowed_file(self, filename: str) -> bool:
        """Check if file extension is allowed."""
        return "." in filename and filename.rsplit(".", 1)[1].lower() in self.ALLOWED_EXTENSIONS

    def _parse_headers(self, worksheet) -> dict:
        """
        Parse Excel headers and map to expected column names.

        Returns:
            Dict mapping expected names to column indices (or empty dict if invalid)
        """
        if not worksheet:
            return {}

        first_row = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        headers = {}

        for idx, cell_value in enumerate(first_row):
            if cell_value is None:
                continue

            # Normalize header: lowercase, strip whitespace
            normalized = str(cell_value).strip().lower().replace(" ", "_")

            # Map to standard column names
            if normalized in ("holiday_date", "date", "holiday", "holiday_date"):
                headers["holiday_date"] = idx
            elif normalized in ("holiday_name", "name", "holiday_name", "holiday"):
                headers["holiday_name"] = idx
            elif normalized in ("holiday_type", "type", "holiday_type"):
                headers["holiday_type"] = idx
            elif normalized in ("description", "desc", "remarks", "notes"):
                headers["description"] = idx

        return headers

    def _parse_row(self, row, headers: dict, row_idx: int) -> dict:
        """
        Parse a single row and extract holiday data.

        Returns:
            Dict with holiday data or {"error": message} if invalid
        """
        try:
            # Extract values from cells
            holiday_date_idx = headers.get("holiday_date")
            holiday_name_idx = headers.get("holiday_name")

            holiday_date_cell = row[holiday_date_idx] if holiday_date_idx is not None else None
            holiday_name_cell = row[holiday_name_idx] if holiday_name_idx is not None else None

            holiday_date_value = holiday_date_cell.value if holiday_date_cell else None
            holiday_name_value = holiday_name_cell.value if holiday_name_cell else None

            # Validate required fields
            if not holiday_date_value or not holiday_name_value:
                return {"error": f"Row {row_idx}: Missing holiday_date or holiday_name"}

            # Parse date
            try:
                if isinstance(holiday_date_value, date):
                    holiday_date = holiday_date_value
                elif isinstance(holiday_date_value, datetime):
                    holiday_date = holiday_date_value.date()
                else:
                    # Try parsing as string
                    parsed = datetime.strptime(str(holiday_date_value).strip(), "%Y-%m-%d").date()
                    holiday_date = parsed
            except (ValueError, TypeError):
                try:
                    # Try alternative date formats
                    parsed = datetime.strptime(str(holiday_date_value).strip(), "%d-%m-%Y").date()
                    holiday_date = parsed
                except (ValueError, TypeError):
                    return {"error": f"Row {row_idx}: Invalid date format '{holiday_date_value}'"}

            # Extract optional fields
            holiday_type_idx = headers.get("holiday_type")
            description_idx = headers.get("description")

            holiday_type = None
            if holiday_type_idx is not None:
                cell = row[holiday_type_idx] if holiday_type_idx < len(row) else None
                holiday_type = cell.value if cell else None
                if holiday_type:
                    holiday_type = str(holiday_type).strip()

            description = None
            if description_idx is not None:
                cell = row[description_idx] if description_idx < len(row) else None
                description = cell.value if cell else None
                if description:
                    description = str(description).strip()

            return {
                "holiday_date": holiday_date,
                "holiday_name": str(holiday_name_value).strip(),
                "holiday_type": holiday_type,
                "description": description,
            }

        except Exception as e:
            logger.error(f"Error parsing row {row_idx}: {e}")
            return {"error": f"Row {row_idx}: {str(e)[:50]}"}
