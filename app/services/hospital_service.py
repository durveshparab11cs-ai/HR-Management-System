"""
app/services/hospital_service.py
==================================
Business logic for Hospital management and Excel imports
"""

import datetime
import os
from typing import Tuple, List, Dict, Optional
from openpyxl import load_workbook
from flask import current_app
from werkzeug.utils import secure_filename

from app.extensions.database import db
from app.repositories.hospital_repository import HospitalRepository
from app.models.hospital import Hospital
from app.models.employee import Employee
from app.models.employee_master import EmployeeMaster


class HospitalService:
    """Service for hospital management operations."""
    
    def __init__(self):
        self.repo = HospitalRepository()
    
    def get_all_hospitals(self, include_inactive: bool = False) -> List[Hospital]:
        """Get all hospitals."""
        return self.repo.get_all(include_inactive)
    
    def get_hospital(self, hospital_id: int) -> Optional[Hospital]:
        """Get hospital by ID."""
        return self.repo.get_by_id(hospital_id)
    
    def search_hospitals(self, query: str) -> List[Hospital]:
        """Search hospitals."""
        return self.repo.search(query)
    
    def create_hospital(
        self,
        hospital_name: str,
        latitude: float,
        longitude: float,
        **kwargs
    ) -> Tuple[bool, str, Optional[Hospital]]:
        """
        Create a new hospital.
        
        Returns:
            (success, message, hospital)
        """
        try:
            # Validation
            if not hospital_name or not hospital_name.strip():
                return False, "Hospital name is required", None
            
            if latitude < -90 or latitude > 90:
                return False, "Invalid latitude (must be between -90 and 90)", None
            
            if longitude < -180 or longitude > 180:
                return False, "Invalid longitude (must be between -180 and 180)", None
            
            # Check for duplicates
            hospital_code = kwargs.get('hospital_code')
            if self.repo.check_duplicate(hospital_name, hospital_code):
                return False, "Hospital with this name or code already exists", None
            
            # Create hospital
            hospital = self.repo.create(
                hospital_name=hospital_name.strip(),
                latitude=latitude,
                longitude=longitude,
                **kwargs
            )
            
            db.session.commit()
            
            current_app.logger.info(f"Hospital created: {hospital.hospital_name} (ID: {hospital.id})")
            
            return True, "Hospital created successfully", hospital
            
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error creating hospital: {str(e)}")
            return False, f"Error: {str(e)}", None
    
    def update_hospital(
        self,
        hospital_id: int,
        **kwargs
    ) -> Tuple[bool, str, Optional[Hospital]]:
        """
        Update hospital details.
        
        Returns:
            (success, message, hospital)
        """
        try:
            hospital = self.repo.get_by_id(hospital_id)
            if not hospital:
                return False, "Hospital not found", None
            
            # Validate coordinates if provided
            if 'latitude' in kwargs:
                lat = kwargs['latitude']
                if lat < -90 or lat > 90:
                    return False, "Invalid latitude", None
            
            if 'longitude' in kwargs:
                lng = kwargs['longitude']
                if lng < -180 or lng > 180:
                    return False, "Invalid longitude", None
            
            # Check duplicate name/code if changed
            new_name = kwargs.get('hospital_name')
            new_code = kwargs.get('hospital_code')
            if new_name or new_code:
                check_name = new_name if new_name else hospital.hospital_name
                check_code = new_code if new_code else hospital.hospital_code
                if self.repo.check_duplicate(check_name, check_code, exclude_id=hospital_id):
                    return False, "Hospital with this name or code already exists", None
            
            # Update
            updated = self.repo.update(hospital_id, **kwargs)
            db.session.commit()
            
            current_app.logger.info(f"Hospital updated: {updated.hospital_name} (ID: {hospital_id})")
            
            return True, "Hospital updated successfully", updated
            
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error updating hospital: {str(e)}")
            return False, f"Error: {str(e)}", None
    
    def delete_hospital(self, hospital_id: int) -> Tuple[bool, str]:
        """
        Delete hospital (soft delete).
        
        Returns:
            (success, message)
        """
        try:
            hospital = self.repo.get_by_id(hospital_id)
            if not hospital:
                return False, "Hospital not found"
            
            # Check if hospital has employees
            employee_count = Employee.query.filter_by(hospital_id=hospital_id).count()
            if employee_count > 0:
                return False, f"Cannot delete hospital. {employee_count} employees are assigned to this hospital."
            
            # Soft delete
            self.repo.delete(hospital_id)
            db.session.commit()
            
            current_app.logger.info(f"Hospital deleted: {hospital.hospital_name} (ID: {hospital_id})")
            
            return True, "Hospital deleted successfully"
            
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error deleting hospital: {str(e)}")
            return False, f"Error: {str(e)}"
    
    def import_hospitals_from_excel(
        self,
        file,
        imported_by_user_id: int
    ) -> Tuple[bool, str, Dict]:
        """
        Import hospitals from Excel file.
        
        Args:
            file: FileStorage object
            imported_by_user_id: User ID who is importing
            
        Returns:
            (success, message, statistics)
        """
        start_time = datetime.datetime.now()
        
        stats = {
            'total_rows': 0,
            'imported': 0,
            'updated': 0,
            'failed': 0,
            'errors': []
        }
        
        try:
            # Save file temporarily
            filename = secure_filename(file.filename)
            temp_path = os.path.join(current_app.config.get('UPLOAD_FOLDER', 'instance/uploads'), filename)
            os.makedirs(os.path.dirname(temp_path), exist_ok=True)
            file.save(temp_path)
            
            # Read Excel using openpyxl
            wb = load_workbook(temp_path, data_only=True)
            ws = wb.active
            
            # Get headers from first row
            headers = {}
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value:
                    headers[cell.value.strip().lower()] = col_idx
            
            stats['total_rows'] = ws.max_row - 1  # Exclude header
            
            current_app.logger.info(f"Processing {stats['total_rows']} hospitals from Excel")
            
            # Process each row
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Extract data using headers
                    row_dict = {}
                    for header_name, col_idx in headers.items():
                        if col_idx <= len(row):
                            row_dict[header_name] = row[col_idx - 1]
                    
                    # Get hospital name
                    hospital_name = str(row_dict.get('hospital  name', '') or '').strip()
                    if not hospital_name or hospital_name == 'nan':
                        stats['errors'].append(f"Row {row_idx}: Hospital name is missing")
                        stats['failed'] += 1
                        continue
                    
                    # Get coordinates
                    try:
                        latitude = float(row_dict.get('latitude', 0) or 0)
                        longitude = float(row_dict.get('longitude', 0) or 0)
                    except (ValueError, TypeError):
                        stats['errors'].append(f"Row {row_idx}: Invalid GPS coordinates")
                        stats['failed'] += 1
                        continue
                    
                    # Validate coordinates
                    if latitude == 0 or longitude == 0:
                        stats['errors'].append(f"Row {row_idx}: {hospital_name} - GPS coordinates missing")
                        stats['failed'] += 1
                        continue
                    
                    if latitude < -90 or latitude > 90 or longitude < -180 or longitude > 180:
                        stats['errors'].append(f"Row {row_idx}: {hospital_name} - Invalid GPS coordinates")
                        stats['failed'] += 1
                        continue
                    
                    # Get other fields
                    location = str(row_dict.get('location', '') or '').strip()
                    if location == 'nan' or not location:
                        location = None
                    
                    status = str(row_dict.get('status', 'Active') or 'Active').strip()
                    is_active = status.lower() == 'active'
                    
                    # Check if hospital exists (by name)
                    existing = self.repo.get_by_name(hospital_name)
                    
                    if existing:
                        # Update existing hospital
                        self.repo.update(
                            existing.id,
                            latitude=latitude,
                            longitude=longitude,
                            location=location,
                            status=status,
                            is_active=is_active
                        )
                        stats['updated'] += 1
                        current_app.logger.info(f"Updated hospital: {hospital_name}")
                    else:
                        # Create new hospital
                        self.repo.create(
                            hospital_name=hospital_name,
                            latitude=latitude,
                            longitude=longitude,
                            location=location,
                            status=status,
                            is_active=is_active,
                            allowed_radius_metres=100  # Default
                        )
                        stats['imported'] += 1
                        current_app.logger.info(f"Imported hospital: {hospital_name}")
                    
                except Exception as e:
                    stats['errors'].append(f"Row {row_idx}: {str(e)}")
                    stats['failed'] += 1
                    continue
            
            # Commit all changes
            db.session.commit()
            
            # Clean up temp file
            try:
                os.remove(temp_path)
            except:
                pass
            
            # Log import
            duration = (datetime.datetime.now() - start_time).total_seconds()
            self._log_import(
                import_type='hospital',
                imported_by=imported_by_user_id,
                filename=filename,
                stats=stats,
                duration=duration
            )
            
            message = f"Import completed: {stats['imported']} created, {stats['updated']} updated, {stats['failed']} failed"
            
            return True, message, stats
            
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error importing hospitals: {str(e)}")
            return False, f"Import failed: {str(e)}", stats
    
    def _log_import(self, import_type: str, imported_by: int, filename: str, stats: Dict, duration: float):
        """Log import operation to database."""
        try:
            from app.models import db as _db
            from sqlalchemy import text
            
            error_log = "\n".join(stats.get('errors', []))[:5000]  # Limit to 5000 chars
            
            _db.session.execute(text("""
                INSERT INTO import_logs (
                    import_type, imported_by, filename,
                    total_rows, rows_imported, rows_updated, rows_failed,
                    hospitals_imported, hospitals_updated,
                    status, error_log, duration_seconds
                ) VALUES (
                    :type, :user, :file,
                    :total, :imported, :updated, :failed,
                    :hosp_imp, :hosp_upd,
                    :status, :errors, :duration
                )
            """), {
                'type': import_type,
                'user': imported_by,
                'file': filename,
                'total': stats.get('total_rows', 0),
                'imported': stats.get('imported', 0),
                'updated': stats.get('updated', 0),
                'failed': stats.get('failed', 0),
                'hosp_imp': stats.get('imported', 0),
                'hosp_upd': stats.get('updated', 0),
                'status': 'completed' if stats.get('failed', 0) == 0 else 'partial',
                'errors': error_log,
                'duration': duration
            })
            
            _db.session.commit()
            
        except Exception as e:
            current_app.logger.error(f"Error logging import: {str(e)}")
