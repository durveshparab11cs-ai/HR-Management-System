"""
blueprints/leave/routes.py
============================
Leave, half-day, and early-leave routes — thin layer only.
"""

from datetime import date
from flask import flash, jsonify, redirect, render_template, request, url_for, send_file
from flask_login import current_user, login_required
import logging

from app.blueprints.employees.repository import EmployeeRepository
from .forms import ApplyEarlyLeaveForm, ApplyHalfDayForm, ApplyLeaveForm, ReviewLeaveForm
from .repository import LeaveRepository
from .service import LeaveService
from . import leave_bp

logger = logging.getLogger(__name__)
_svc  = LeaveService()
_repo = LeaveRepository()
_emp  = EmployeeRepository()


def _get_employee_or_redirect():
    emp = _emp.get_by_user_id(current_user.id)
    if not emp:
        flash("Employee profile not found. Contact HR.", "warning")
        return None
    return emp


def _get_manager_code_by_name(manager_name: str) -> str:
    """Look up manager's employee code from employee_master by name."""
    from app.models.employee_master import EmployeeMaster  # noqa: PLC0415
    if not manager_name:
        return ""
    master = EmployeeMaster.query.filter_by(
        employee_name=manager_name,
        is_active=True
    ).first()
    return master.employee_code if master else ""


# ── Manager Code Lookup (AJAX) ────────────────────────────────────────

@leave_bp.route("/lookup-manager")
@login_required
def lookup_manager():
    """AJAX: validate a reporting manager code and return their details."""
    from app.models.employee_master import EmployeeMaster  # noqa: PLC0415
    code = request.args.get("code", "").strip().upper()
    if not code:
        return jsonify(found=False, message="Enter an Employee Code.")
    my_emp = _emp.get_by_user_id(current_user.id)
    if my_emp and my_emp.employee_code.upper() == code:
        return jsonify(found=False, message="You cannot select yourself as Reporting Manager.")
    master = EmployeeMaster.query.filter_by(employee_code=code, is_active=True).first()
    if not master:
        return jsonify(found=False, message="Reporting Manager not found.")
    return jsonify(
        found=True,
        name=master.employee_name,
        department=master.department or "—",
        designation=master.designation or "—",
    )


# ── Manager: Leave Approval Dashboard ────────────────────────────────

@leave_bp.route("/my-approvals")
@login_required
def my_approvals():
    """Requests assigned to the logged-in employee as reporting manager."""
    emp = _get_employee_or_redirect()
    if not emp:
        return redirect(url_for("dashboard.index"))
    
    # Get manager's name from employee_master
    from app.models.employee_master import EmployeeMaster  # noqa: PLC0415
    manager_master = EmployeeMaster.query.filter_by(
        employee_code=emp.employee_code.upper(),
        is_active=True
    ).first()
    
    if not manager_master:
        flash("Your employee profile not found in master data.", "warning")
        return redirect(url_for("dashboard.index"))
    
    mgr_name = manager_master.employee_name
    mgr_code = emp.employee_code.upper()
    
    status_filter = request.args.get("status", "")
    page = request.args.get("page", 1, type=int)

    # Safely query — filter by both name and code for backward compatibility
    lr_list = []
    hd_list = []
    el_list = []
    hd_pag  = type('P', (), {'items': []})()
    el_pag  = type('P', (), {'items': []})()

    try:
        from app.models.leave import LeaveRequest  # noqa: PLC0415
        from sqlalchemy import or_  # noqa: PLC0415
        lr_q = LeaveRequest.query.filter(
            or_(
                LeaveRequest.reporting_manager_name == mgr_name,
                LeaveRequest.reporting_manager_code == mgr_code
            ),
            LeaveRequest.is_deleted == False
        )
        if status_filter:
            lr_q = lr_q.filter_by(status=status_filter)
        lr_list = lr_q.order_by(LeaveRequest.applied_on.desc()).limit(30).all()
    except Exception:  # noqa: BLE001
        lr_list = []

    try:
        hd_pag = _repo.get_halfdays_for_manager_by_name(mgr_name, page=page, status=status_filter)
        hd_list = hd_pag.items
    except Exception:  # noqa: BLE001
        # Fallback to code-based filter
        try:
            hd_pag = _repo.get_halfdays_for_manager(mgr_code, page=page, status=status_filter)
            hd_list = hd_pag.items
        except Exception:  # noqa: BLE001
            hd_list = []

    try:
        el_pag = _repo.get_earlyleaves_for_manager_by_name(mgr_name, page=page, status=status_filter)
        el_list = el_pag.items
    except Exception:  # noqa: BLE001
        # Fallback to code-based filter
        try:
            el_pag = _repo.get_earlyleaves_for_manager(mgr_code, page=page, status=status_filter)
            el_list = el_pag.items
        except Exception:  # noqa: BLE001
            el_list = []

    return render_template(
        "leave/my_approvals.html",
        title="Leave Approval",
        employee=emp,
        lr_list=lr_list,
        hd_list=hd_list,
        el_list=el_list,
        hd_pag=hd_pag,
        el_pag=el_pag,
        status_filter=status_filter,
    )


# ─── Leave Portal Index ──────────────────────────────────────────────

@leave_bp.route("/")
@login_required
def index():
    emp = _get_employee_or_redirect()
    if not emp:
        return redirect(url_for("dashboard.index"))
    balances = _svc.get_balance(emp.id)
    page = request.args.get("page", 1, type=int)
    pagination = _repo.get_employee_requests(emp.id, page=page)
    hd_pagination  = _repo.get_employee_halfdays(emp.id, page=1, per_page=5)
    el_pagination  = _repo.get_employee_earlyleaves(emp.id, page=1, per_page=5)
    return render_template(
        "leave/index.html", title="Leave Portal",
        employee=emp, balances=balances,
        pagination=pagination,
        hd_list=hd_pagination.items,
        el_list=el_pagination.items,
    )


# ─── Apply Leave ─────────────────────────────────────────────────────

@leave_bp.route("/apply", methods=["GET", "POST"])
@login_required
def apply():
    from .forms import get_manager_choices  # noqa: PLC0415
    emp = _get_employee_or_redirect()
    if not emp: return redirect(url_for("dashboard.index"))
    form = ApplyLeaveForm()
    
    # Filter to only show the 4 approved leave types (CL, SL, PL, CO)
    allowed_codes = ['CL', 'SL', 'PL', 'CO']
    all_types = _repo.get_all_types()
    filtered_types = [lt for lt in all_types if lt.code in allowed_codes]
    # Sort by leave_order or id for consistent display
    filtered_types.sort(key=lambda x: (x.leave_order or 0, x.id))
    form.leave_type_id.choices = [(lt.id, lt.name) for lt in filtered_types]
    form.reporting_manager.choices = get_manager_choices()
    
    if form.validate_on_submit():
        att = request.files.get("attachment")
        # Get manager code from employee_master if exists
        manager_name = form.reporting_manager.data
        manager_code = _get_manager_code_by_name(manager_name)
        
        ok, msg, lr = _svc.apply_leave(
            employee_id=emp.id,
            form_data={
                "start_date": form.start_date.data,
                "end_date": form.end_date.data,
                "leave_type_id": form.leave_type_id.data,
                "reason": form.reason.data,
                "reporting_manager_name": manager_name,
                "reporting_manager_code": manager_code or "",
            },
            attachment=att if (att and att.filename) else None,
        )
        flash(msg, "success" if ok else "danger")
        if ok:
            return redirect(url_for("leave.index"))
    return render_template("leave/apply.html", title="Apply for Leave", form=form)


# ─── Cancel Leave ────────────────────────────────────────────────────

@leave_bp.route("/<int:lr_id>/cancel", methods=["POST"])
@login_required
def cancel(lr_id: int):
    emp = _get_employee_or_redirect()
    if not emp: return redirect(url_for("dashboard.index"))
    ok, msg = _svc.cancel_leave(lr_id, emp.id)
    flash(msg, "success" if ok else "danger")
    return redirect(url_for("leave.index"))


# ─── Manager/HR: Review Leave ────────────────────────────────────────

@leave_bp.route("/pending")
@login_required
def pending():
    from app.core.dept_filter import get_dept_filter  # noqa: PLC0415
    page = request.args.get("page", 1, type=int)
    try:
        dept_filter = get_dept_filter()
    except Exception:  # noqa: BLE001
        dept_filter = None
    try:
        pagination = _repo.get_pending(page=page, department=dept_filter)
    except TypeError:
        pagination = _repo.get_pending(page=page)
    try:
        hd_pag = _repo.get_pending_halfdays(page=1, per_page=10, department=dept_filter)
    except TypeError:
        hd_pag = _repo.get_pending_halfdays(page=1, per_page=10)
    try:
        el_pag = _repo.get_pending_earlyleaves(page=1, per_page=10, department=dept_filter)
    except TypeError:
        el_pag = _repo.get_pending_earlyleaves(page=1, per_page=10)
    return render_template(
        "leave/pending.html", title="Pending Approvals",
        pagination=pagination,
        hd_list=hd_pag.items,
        el_list=el_pag.items,
    )


@leave_bp.route("/<int:lr_id>/approve", methods=["POST"])
@login_required
def approve(lr_id: int):
    form = ReviewLeaveForm()
    ok, msg = _svc.approve_leave(lr_id, current_user.id, form.comment.data or "")
    flash(msg, "success" if ok else "danger")
    return redirect(request.referrer or url_for("leave.pending"))


@leave_bp.route("/<int:lr_id>/reject", methods=["POST"])
@login_required
def reject(lr_id: int):
    form = ReviewLeaveForm()
    comment = form.comment.data or ""
    
    # Validate that rejection reason is provided
    if not comment or comment.strip() == "":
        flash("Rejection reason is mandatory. Please provide a reason for rejection.", "danger")
        return redirect(request.referrer or url_for("leave.pending"))
    
    ok, msg = _svc.reject_leave(lr_id, current_user.id, comment)
    flash(msg, "success" if ok else "danger")
    return redirect(request.referrer or url_for("leave.pending"))


# ─── Half Day ────────────────────────────────────────────────────────

@leave_bp.route("/halfday/apply", methods=["GET", "POST"])
@login_required
def apply_halfday():
    from .forms import get_manager_choices  # noqa: PLC0415
    emp = _get_employee_or_redirect()
    if not emp: return redirect(url_for("dashboard.index"))
    form = ApplyHalfDayForm()
    form.reporting_manager.choices = get_manager_choices()
    
    if form.validate_on_submit():
        manager_name = form.reporting_manager.data
        manager_code = _get_manager_code_by_name(manager_name)
        
        ok, msg, _ = _svc.apply_halfday(emp.id, {
            "date": form.date.data,
            "half_type": form.half_type.data,
            "reason": form.reason.data,
            "reporting_manager_name": manager_name,
            "reporting_manager_code": manager_code or "",
        })
        flash(msg, "success" if ok else "danger")
        if ok: return redirect(url_for("leave.index"))
    return render_template("leave/apply_halfday.html", title="Request Half Day", form=form)


@leave_bp.route("/halfday/<int:hd_id>/approve", methods=["POST"])
@login_required
def approve_halfday(hd_id: int):
    ok, msg = _svc.approve_halfday(hd_id, current_user.id)
    flash(msg, "success" if ok else "danger")
    return redirect(request.referrer or url_for("leave.pending"))


@leave_bp.route("/halfday/<int:hd_id>/reject", methods=["POST"])
@login_required
def reject_halfday(hd_id: int):
    form = ReviewLeaveForm()
    comment = form.comment.data or ""
    
    # Validate that rejection reason is provided
    if not comment or comment.strip() == "":
        flash("Rejection reason is mandatory. Please provide a reason for rejection.", "danger")
        return redirect(request.referrer or url_for("leave.pending"))
    
    ok, msg = _svc.reject_halfday(hd_id, current_user.id)
    flash(msg, "success" if ok else "danger")
    return redirect(request.referrer or url_for("leave.pending"))


# ─── Early Leave ─────────────────────────────────────────────────────

@leave_bp.route("/earlyleave/apply", methods=["GET", "POST"])
@login_required
def apply_earlyleave():
    from .forms import get_manager_choices  # noqa: PLC0415
    emp = _get_employee_or_redirect()
    if not emp: return redirect(url_for("dashboard.index"))
    form = ApplyEarlyLeaveForm()
    form.reporting_manager.choices = get_manager_choices()
    
    if form.validate_on_submit():
        manager_name = form.reporting_manager.data
        manager_code = _get_manager_code_by_name(manager_name)
        
        ok, msg, _ = _svc.apply_earlyleave(emp.id, {
            "date": form.date.data,
            "requested_leave_time": form.requested_leave_time.data,
            "reason": form.reason.data,
            "reporting_manager_name": manager_name,
            "reporting_manager_code": manager_code or "",
        })
        flash(msg, "success" if ok else "danger")
        if ok: return redirect(url_for("leave.index"))
    return render_template("leave/apply_earlyleave.html", title="Request Early Leave", form=form)


@leave_bp.route("/earlyleave/<int:el_id>/approve", methods=["POST"])
@login_required
def approve_earlyleave(el_id: int):
    ok, msg = _svc.approve_earlyleave(el_id, current_user.id)
    flash(msg, "success" if ok else "danger")
    return redirect(request.referrer or url_for("leave.pending"))


@leave_bp.route("/earlyleave/<int:el_id>/reject", methods=["POST"])
@login_required
def reject_earlyleave(el_id: int):
    form = ReviewLeaveForm()
    comment = form.comment.data or ""
    
    # Validate that rejection reason is provided
    if not comment or comment.strip() == "":
        flash("Rejection reason is mandatory. Please provide a reason for rejection.", "danger")
        return redirect(request.referrer or url_for("leave.pending"))
    
    ok, msg = _svc.reject_earlyleave(el_id, current_user.id)
    flash(msg, "success" if ok else "danger")
    return redirect(request.referrer or url_for("leave.pending"))


# ─── Comp Off Management ─────────────────────────────────────────────

# ─── Export Leave Requests to Excel ──────────────────────────────────

@leave_bp.route("/export")
@login_required
def export_leave_requests():
    """Export all leave requests (for current employee or all if admin) to Excel."""
    from io import BytesIO  # noqa: PLC0415
    from openpyxl import Workbook  # noqa: PLC0415
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # noqa: PLC0415
    from openpyxl.utils import get_column_letter  # noqa: PLC0415
    
    emp = _emp.get_by_user_id(current_user.id)
    if not emp:
        flash("Employee profile not found.", "warning")
        return redirect(url_for("leave.index"))
    
    # Get leave requests - all for admin, only own for employee
    if current_user.role in ("super_admin", "admin", "hr_manager", "hr_staff"):
        # Admin/HR can export all leave requests
        leave_requests = _repo.get_all_requests_no_pagination()
    else:
        # Regular employee - only their own
        leave_requests = _repo.get_employee_requests_all(emp.id)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Leave Requests"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Define headers
    headers = [
        "Emp Code", "Employee Name", "Leave Type", "From Date", "To Date", 
        "Total Days", "Reason", "Status", "Applied On", "Reviewed By", 
        "Reviewed On", "Reviewer Comment", "Manager Code", "Manager Name"
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Write data rows
    row_num = 2
    for lr in leave_requests:
        try:
            # Get employee info
            emp_code = lr.employee.employee_code if lr.employee else "N/A"
            emp_name = lr.employee.full_name if lr.employee else "N/A"
            leave_type_name = lr.leave_type.name if lr.leave_type else "N/A"
            
            # Get reviewer name
            reviewer_name = lr.reviewer.full_name if lr.reviewer else "—"
            
            # Prepare row data
            row_data = [
                emp_code,
                emp_name,
                leave_type_name,
                lr.start_date.strftime("%d-%m-%Y") if lr.start_date else "",
                lr.end_date.strftime("%d-%m-%Y") if lr.end_date else "",
                lr.total_days,
                lr.reason or "—",
                lr.status.upper(),
                lr.applied_on.strftime("%d-%m-%Y %H:%M") if lr.applied_on else "",
                reviewer_name,
                lr.reviewed_on.strftime("%d-%m-%Y %H:%M") if lr.reviewed_on else "—",
                lr.reviewer_comment or "—",
                lr.reporting_manager_code or "—",
                lr.reporting_manager_name or "—",
            ]
            
            # Write row
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
                
                # Apply alignment based on column type
                if col_num in (1, 6, 8):  # Numbers and status
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
                
                # Color code status
                if col_num == 8:  # Status column
                    if value == "APPROVED":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value == "REJECTED":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    elif value == "PENDING":
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            
            row_num += 1
        except Exception as e:
            logger.warning(f"Error processing leave request {lr.id}: {e}")
            continue
    
    # Adjust column widths
    column_widths = [12, 20, 18, 12, 12, 11, 25, 12, 18, 18, 18, 20, 12, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Create response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    from datetime import datetime  # noqa: PLC0415
    filename = f"Leave_Requests_{datetime.now().strftime('%d-%m-%Y_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@leave_bp.route("/comp-off/status")
@login_required
def comp_off_status():
    """
    Employee view: Check available comp offs and expiry information.
    AJAX endpoint returning JSON.
    """
    emp = _get_employee_or_redirect()
    if not emp:
        return jsonify(error="Employee not found"), 404
    
    from .comp_off_service import CompOffService  # noqa: PLC0415
    comp_svc = CompOffService()
    
    available = comp_svc.get_available_comp_offs(emp.id)
    expiry_info = comp_svc.check_expired_comp_offs(emp.id)
    
    return jsonify({
        "available_count": len(available),
        "available_comp_offs": [
            {
                "id": co.id,
                "work_date": co.comp_off_work_date.isoformat() if co.comp_off_work_date else None,
                "expiry_date": co.comp_off_expiry_date.isoformat() if co.comp_off_expiry_date else None,
                "days_left": (co.comp_off_expiry_date - date.today()).days if co.comp_off_expiry_date else 0,
            }
            for co in available
        ],
        "expiry_info": expiry_info,
    })


@leave_bp.route("/admin/comp-off/earn", methods=["POST"])
@login_required
def admin_earn_comp_off():
    """
    Admin endpoint: Mark that an employee worked on holiday and earned comp off.
    POST: {employee_id, work_date, holiday_name}
    """
    from flask import current_user  # noqa: PLC0415
    from app.models.user import User  # noqa: PLC0415
    from .comp_off_service import CompOffService  # noqa: PLC0415
    from datetime import datetime  # noqa: PLC0415
    
    # Check authorization - only admin/HR can do this
    user = User.query.get(current_user.id)
    if user.role not in ["admin", "hr_manager", "hr_staff"]:
        return jsonify(error="Unauthorized"), 403
    
    data = request.get_json() or {}
    emp_id = data.get("employee_id")
    work_date_str = data.get("work_date")
    holiday_name = data.get("holiday_name", "")
    
    if not emp_id or not work_date_str:
        return jsonify(error="Missing required fields"), 400
    
    try:
        work_date = datetime.fromisoformat(work_date_str).date()
    except (ValueError, TypeError):
        return jsonify(error="Invalid work_date format"), 400
    
    comp_svc = CompOffService()
    ok, msg = comp_svc.earn_comp_off(emp_id, work_date, holiday_name)
    
    return jsonify(success=ok, message=msg), (200 if ok else 400)


@leave_bp.route("/admin/comp-off/list")
@login_required
def admin_comp_off_list():
    """
    Admin endpoint: List all comp offs (earned, used, expired).
    Query params: status (earned|used|expired), employee_id (optional)
    """
    from flask import current_user  # noqa: PLC0415
    from app.models.user import User  # noqa: PLC0415
    from app.models.leave import LeaveRequest  # noqa: PLC0415
    from datetime import timedelta  # noqa: PLC0415
    
    # Check authorization
    user = User.query.get(current_user.id)
    if user.role not in ["admin", "hr_manager", "hr_staff"]:
        return jsonify(error="Unauthorized"), 403
    
    status_filter = request.args.get("status", "earned")  # earned|used|expired
    emp_id = request.args.get("employee_id", type=int)
    page = request.args.get("page", 1, type=int)
    per_page = 20
    
    today = date.today()
    query = LeaveRequest.query.filter(
        LeaveRequest.leave_type.has(code='CO'),
        LeaveRequest.is_deleted == False,
    )
    
    if emp_id:
        query = query.filter(LeaveRequest.employee_id == emp_id)
    
    if status_filter == "earned":
        query = query.filter(
            LeaveRequest.comp_off_expiry_date >= today,
            LeaveRequest.comp_off_used_on == None,
        )
    elif status_filter == "used":
        query = query.filter(LeaveRequest.comp_off_used_on != None)
    elif status_filter == "expired":
        query = query.filter(
            LeaveRequest.comp_off_expiry_date < today,
            LeaveRequest.comp_off_used_on == None,
        )
    
    pagination = query.paginate(page=page, per_page=per_page)
    
    return jsonify({
        "total": pagination.total,
        "page": page,
        "pages": pagination.pages,
        "comp_offs": [
            {
                "id": co.id,
                "employee_id": co.employee_id,
                "employee_code": co.employee.employee_code if co.employee else "N/A",
                "employee_name": co.employee.full_name if co.employee else "N/A",
                "work_date": co.comp_off_work_date.isoformat() if co.comp_off_work_date else None,
                "expiry_date": co.comp_off_expiry_date.isoformat() if co.comp_off_expiry_date else None,
                "used_on": co.comp_off_used_on.isoformat() if co.comp_off_used_on else None,
                "status": "used" if co.comp_off_used_on else ("expired" if co.comp_off_expiry_date < today else "available"),
            }
            for co in pagination.items
        ]
    })

