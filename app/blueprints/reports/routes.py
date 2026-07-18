"""blueprints/reports/routes.py"""

from datetime import date, datetime
from flask import flash, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required

from app.core.security import manager_required
from app.constants.limits import Limits
from app.extensions.limiter import limiter
from .service import ReportService
from . import reports_bp

_svc = ReportService()


@reports_bp.route("/")
@reports_bp.route("")
@login_required
@manager_required
def index():
    stats = _svc.employee_stats()
    today = date.today()
    daily = _svc.daily_attendance(today)
    return render_template("reports/index.html", title="Reports", stats=stats, daily=daily, today=today)


@reports_bp.route("/attendance")
@login_required
@manager_required
def attendance_report():
    start_str  = request.args.get("start", date.today().replace(day=1).isoformat())
    end_str    = request.args.get("end",   date.today().isoformat())
    department = request.args.get("dept",  "")
    view       = request.args.get("view",  "datewise")
    try:
        start = date.fromisoformat(start_str)
        end   = date.fromisoformat(end_str)
    except ValueError:
        start = date.today().replace(day=1)
        end   = date.today()

    summary_data  = _svc.attendance_summary(start, end, department=department)
    datewise_flat = _svc.attendance_datewise(start, end, department=department)

    # Group flat list by date for clean Jinja rendering
    from collections import OrderedDict  # noqa: PLC0415
    grouped = OrderedDict()
    for rec in datewise_flat:
        key = rec["date"]
        if key not in grouped:
            grouped[key] = {"day": rec["day"], "records": []}
        grouped[key]["records"].append(rec)

    depts = _svc.get_departments()
    return render_template(
        "reports/attendance.html", title="Attendance Report",
        data=summary_data,
        grouped=grouped,
        start=start, end=end,
        department=department,
        departments=depts,
        view=view,
    )


@reports_bp.route("/leave")
@login_required
@manager_required
def leave_report():
    year = request.args.get("year", date.today().year, type=int)
    data = _svc.leave_summary(year)
    return render_template("reports/leave.html", title="Leave Report", data=data, year=year)


@reports_bp.route("/employees")
@login_required
@manager_required
def employee_report():
    stats    = _svc.employee_stats()
    employees = _svc.get_all_employees_simple()
    return render_template("reports/employees.html", title="Employee Report", stats=stats, employees=employees)


@reports_bp.route("/export/attendance-datewise")
@login_required
@manager_required
@limiter.limit(Limits.RateLimit.EXPORT)
def export_attendance_datewise():
    """Export date-wise attendance as a formatted Excel file."""
    import io  # noqa: PLC0415
    import pytz  # noqa: PLC0415
    from openpyxl import Workbook  # noqa: PLC0415
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # noqa: PLC0415
    from flask import Response  # noqa: PLC0415

    start_str  = request.args.get("start", date.today().replace(day=1).isoformat())
    end_str    = request.args.get("end",   date.today().isoformat())
    department = request.args.get("dept", "")
    try:
        start = date.fromisoformat(start_str)
        end   = date.fromisoformat(end_str)
    except ValueError:
        start = date.today().replace(day=1)
        end   = date.today()

    data = _svc.attendance_datewise(start, end, department=department)

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # ── Styles ──────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="1A3C6E")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    subhdr_fill = PatternFill("solid", fgColor="D6E4F0")
    subhdr_font = Font(bold=True, size=10)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left",   vertical="center")
    thin        = Side(style="thin", color="CCCCCC")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    status_colors = {
        "Present":   "E8F5E9",
        "Absent":    "FFEBEE",
        "Half Day":  "FFF8E1",
        "On Leave":  "E3F2FD",
        "Holiday":   "F3E5F5",
        "Weekend":   "F5F5F5",
    }

    # ── Title row ────────────────────────────────────────────────────
    dept_label = f" — {department}" if department else ""
    ws.merge_cells("A1:O1")
    ws["A1"].value       = f"Attendance Report{dept_label}  |  {start.strftime('%d %b %Y')} to {end.strftime('%d %b %Y')}"
    ws["A1"].font        = Font(bold=True, size=13, color="1A3C6E")
    ws["A1"].alignment   = center
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:O2")
    ws["A2"].value     = f"Generated: {datetime.now(pytz.timezone('Asia/Kolkata')).strftime('%d %b %Y %I:%M %p IST')}  |  Total Records: {len(data)}"
    ws["A2"].font      = Font(italic=True, size=9, color="666666")
    ws["A2"].alignment = center
    ws.row_dimensions[2].height = 18

    # ── Column headers ────────────────────────────────────────────────
    headers = [
        "Date", "Day", "Emp Code", "Employee Name", "Department",
        "Designation", "Check In (IST)", "Check Out (IST)",
        "Working Hours", "Status", "Late", "Late (min)",
        "Overtime (min)", "GPS Location", "Remarks",
    ]
    col_widths = [14, 12, 12, 24, 18, 20, 14, 14, 14, 12, 8, 12, 14, 28, 18]

    for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
        cell                = ws.cell(row=3, column=col_idx, value=hdr)
        cell.font           = header_font
        cell.fill           = header_fill
        cell.alignment      = center
        cell.border         = border
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[3].height = 20

    # Freeze header rows
    ws.freeze_panes = "A4"

    # ── Data rows ─────────────────────────────────────────────────────
    prev_date = None
    for row_idx, rec in enumerate(data, start=4):
        # Subtle date-group separator — light blue row when date changes
        is_new_date = rec["date"] != prev_date
        prev_date = rec["date"]

        row_fill = None
        status_key = rec["status"]
        if status_key in status_colors:
            row_fill = PatternFill("solid", fgColor=status_colors[status_key])

        values = [
            rec["date"],
            rec["day"],
            rec["emp_code"],
            rec["emp_name"],
            rec["department"],
            rec["designation"],
            rec["check_in"],
            rec["check_out"],
            f'{rec["working_hours"]:.2f}',
            rec["status"],
            rec["is_late"],
            rec["late_minutes"],
            rec["overtime_min"],
            rec["location"],
            "",  # Remarks — blank for HR to fill
        ]

        for col_idx, val in enumerate(values, start=1):
            cell            = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border     = border
            cell.alignment  = center if col_idx not in (4, 5, 6, 14) else left
            if row_fill:
                cell.fill = row_fill
            # Bold date column when date changes
            if col_idx <= 2 and is_new_date:
                cell.font = Font(bold=True, size=9)
            else:
                cell.font = Font(size=9)

        # Late marker — orange text for late rows
        if rec["is_late"] == "Yes":
            ws.cell(row=row_idx, column=11).font = Font(color="E65100", bold=True, size=9)
            ws.cell(row=row_idx, column=12).font = Font(color="E65100", bold=True, size=9)

        ws.row_dimensions[row_idx].height = 16

    # ── Summary section ───────────────────────────────────────────────
    if data:
        summary_row = len(data) + 5
        ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True, size=11)

        total_present = sum(1 for r in data if "Present" in r["status"])
        total_absent  = sum(1 for r in data if "Absent"  in r["status"])
        total_half    = sum(1 for r in data if "Half"    in r["status"])
        total_late    = sum(1 for r in data if r["is_late"] == "Yes")
        unique_emp    = len({r["emp_code"] for r in data})

        summary_items = [
            ("Total Employees",   unique_emp),
            ("Total Records",     len(data)),
            ("Present Days",      total_present),
            ("Absent Days",       total_absent),
            ("Half Days",         total_half),
            ("Late Arrivals",     total_late),
        ]
        for i, (label, val) in enumerate(summary_items):
            r = summary_row + i + 1
            ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=9)
            ws.cell(row=r, column=2, value=val).font   = Font(size=9)

    # ── Output ────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname   = f"Attendance_DateWise_{start.strftime('%d%b%Y')}_to_{end.strftime('%d%b%Y')}.xlsx"
    dept_fn = f"_{department.replace(' ','_')}" if department else ""
    fname   = f"Attendance_DateWise{dept_fn}_{start.strftime('%d%b%Y')}_to_{end.strftime('%d%b%Y')}.xlsx"

    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@reports_bp.route("/export/<report_type>/<fmt>")
@login_required
@manager_required
@limiter.limit(Limits.RateLimit.EXPORT)
def export(report_type: str, fmt: str):
    """Export report as CSV, Excel or PDF."""
    from app.utils.export_utils import export_to_csv, export_to_excel, export_to_pdf, stream_file_response
    start_str = request.args.get("start", date.today().replace(day=1).isoformat())
    end_str   = request.args.get("end",   date.today().isoformat())
    try:
        start = date.fromisoformat(start_str)
        end   = date.fromisoformat(end_str)
    except ValueError:
        start = date.today().replace(day=1)
        end   = date.today()

    if report_type == "attendance":
        data    = _svc.attendance_summary(start, end)
        headers = ["Code","Name","Department","Present","Absent","Half Day","On Leave","Late Days","Total Hours"]
        rows    = [
            [r["employee"].employee_code, r["user"].full_name,
             r["employee"].department or "—",
             r["present"], r["absent"], r["half_day"],
             r["on_leave"], r["late_days"], r["total_hours"]]
            for r in data
        ]
        title = f"Attendance Report ({start} to {end})"
    elif report_type == "leave":
        year = request.args.get("year", date.today().year, type=int)
        data    = _svc.leave_summary(year)
        headers = ["Code","Name","Total Requests","Approved Days","Pending","Rejected"]
        rows    = [
            [r["employee"].employee_code, r["user"].full_name,
             r["total_requests"], r["approved_days"], r["pending_count"], r["rejected_count"]]
            for r in data
        ]
        title = f"Leave Report {year}"
    elif report_type == "employees":
        data    = _svc.get_all_employees_simple()
        headers = ["Code","Name","Email","Department","Designation","Employment Type"]
        rows    = [
            [e.employee_code, u.full_name, u.email,
             e.department or "—", e.designation or "—",
             e.employment_type.replace("_"," ").title()]
            for e, u in data
        ]
        title = "Employee Report"
    else:
        flash("Unknown report type.", "danger")
        return redirect(url_for("reports.index"))

    if fmt == "csv":
        file_bytes, mime, fname = export_to_csv(headers, rows, report_type)
    elif fmt == "xlsx":
        file_bytes, mime, fname = export_to_excel(headers, rows, sheet_name=title[:31], filename=report_type)
    elif fmt == "pdf":
        file_bytes, mime, fname = export_to_pdf(title, headers, rows, filename=report_type)
    else:
        flash("Invalid export format.", "danger")
        return redirect(url_for("reports.index"))

    return stream_file_response(file_bytes, mime, fname)
