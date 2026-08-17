"""
attendance/routes.py
======================
Attendance routes — thin HTTP layer only.
"""

import os
import traceback
import logging
from datetime import date
from flask import flash, jsonify, redirect, render_template, request, url_for, send_from_directory, abort
from flask_login import current_user, login_required

from app.blueprints.employees.repository import EmployeeRepository
from app.core.security import hr_required
from app.repositories.office_device_repo import OfficeDeviceRepository
from .device_service import DeviceService
from .forms import OfficeSettingsForm
from .office_settings_service import OfficeSettingsService
from .repository import AttendanceRepository
from .service import AttendanceService
from . import attendance_bp

logger = logging.getLogger("attendance")

_svc       = AttendanceService()
_repo      = AttendanceRepository()
_emp_repo  = EmployeeRepository()
_office_svc = OfficeSettingsService()
_device_repo = OfficeDeviceRepository()
_device_svc = DeviceService(_device_repo)


# ── Employee attendance dashboard ────────────────────────────────────

@attendance_bp.route("/gps-test")
@login_required
def gps_test():
    """GPS diagnostic page — helps identify why geolocation is blocked."""
    return render_template("attendance/gps_test.html", title="GPS Test")


@attendance_bp.route("/")
@login_required
def index():
    employee = _emp_repo.get_by_user_id(current_user.id)
    if not employee:
        flash("Employee profile not found. Contact HR.", "warning")
        return redirect(url_for("dashboard.index"))

    status = _svc.get_today_status(employee)
    page   = request.args.get("page", 1, type=int)
    history = _repo.get_history_with_photos(employee.id, page=page, per_page=30)
    office = status.get("office")

    return render_template(
        "attendance/dashboard.html",
        title="Attendance",
        employee=employee,
        status=status,
        history=history,
        office=office,
    )


# ── Check In (AJAX) ──────────────────────────────────────────────────

@attendance_bp.route("/checkin", methods=["POST"])
@login_required
def checkin():
    """
    Check-in endpoint — validates GPS + photo, creates attendance record.
    """
    logger.info("===== CHECK IN START =====")
    logger.info("User ID: %s", current_user.id)
    logger.info("User Email: %s", current_user.email)
    
    try:
        employee = _emp_repo.get_by_user_id(current_user.id)
        if not employee:
            logger.error("CHECK IN FAILED: Employee profile not found for user_id=%s", current_user.id)
            return jsonify(success=False, message="Employee profile not found."), 400
        
        logger.info("Employee ID: %s", employee.id)
        logger.info("Employee Code: %s", employee.employee_code)
        logger.info("Employee Name: %s", employee.full_name)

        # DEVICE VALIDATION — check if request is from allowed office computer
        if _device_svc.is_device_check_enabled():
            is_allowed, device_msg = _device_svc.is_allowed_device()
            logger.info("Device validation: allowed=%s, ip=%s", is_allowed, _device_svc.get_client_ip())
            if not is_allowed:
                logger.error("CHECK IN FAILED: %s", device_msg)
                return jsonify(success=False, message=device_msg), 403  # 403 Forbidden

        # MANDATORY PHOTO VALIDATION — check if photo was uploaded
        from app.models.attendance_photo import AttendancePhoto  # noqa: PLC0415
        today = date.today()
        logger.info("Today's date: %s", today)
        
        # Get today's attendance to check for photo
        attendance_today = _repo.get_today(employee.id, today)
        logger.info("Existing attendance today: %s", attendance_today)
        
        if attendance_today and attendance_today.id:
            # Check if photo exists for this attendance
            photo = AttendancePhoto.query.filter_by(
                attendance_id=attendance_today.id
            ).first()
            logger.info("Photo record found: %s", photo)
            
            if not photo or (not photo.image_data and not photo.file_path):
                logger.error("CHECK IN FAILED: No photo found for attendance_id=%s", attendance_today.id)
                return jsonify(
                    success=False,
                    message="⚠️ Proof Photo is required to mark attendance. Please upload your GPS Map Camera selfie first."
                ), 400
            
            logger.info("Photo validation PASSED - image_data exists: %s, file_path: %s", 
                       bool(photo.image_data), photo.file_path)
        else:
            logger.error("CHECK IN FAILED: No attendance record exists yet (photo should have created one)")
            return jsonify(
                success=False,
                message="⚠️ Proof Photo is required to mark attendance. Please upload your GPS Map Camera selfie first."
            ), 400

        lat = request.form.get("latitude", "")
        lon = request.form.get("longitude", "")
        acc = request.form.get("accuracy", "")
        
        logger.info("GPS Data - Lat: %s, Lon: %s, Accuracy: %s", lat, lon, acc)

        ok, message, attendance, gps_detail = _svc.check_in(employee, lat, lon, acc)
        
        logger.info("Service check_in result: ok=%s, message=%s", ok, message)
        logger.info("GPS Detail: %s", gps_detail)
        
        if ok:
            logger.info("CHECK IN SUCCESS: attendance_id=%s, time=%s", 
                       attendance.id, attendance.check_in_time)
            logger.info("===== CHECK IN END (SUCCESS) =====")
            
            # Ensure check_in_time is a datetime object before formatting
            checkin_time_str = "—"
            if attendance.check_in_time:
                try:
                    if hasattr(attendance.check_in_time, 'strftime'):
                        checkin_time_str = attendance.check_in_time.strftime("%H:%M")
                    else:
                        # Fallback if somehow it's a string
                        checkin_time_str = str(attendance.check_in_time)[-5:] if isinstance(attendance.check_in_time, str) else "—"
                except Exception as fmt_err:
                    logger.error("TIME FORMATTING ERROR: %s", str(fmt_err))
                    checkin_time_str = "—"
            
            return jsonify(
                success=True,
                message=message,
                time=checkin_time_str,
                is_late=attendance.is_late,
                late_minutes=attendance.late_minutes or 0,
                gps=gps_detail,
            )
        
        logger.error("CHECK IN FAILED: %s", message)
        logger.info("===== CHECK IN END (FAILED) =====")
        return jsonify(success=False, message=message, gps=gps_detail), 400
        
    except AttributeError as ae:
        logger.error("===== CHECK IN ATTRIBUTE ERROR =====")
        logger.error("Missing attribute: %s", str(ae))
        import traceback
        logger.error("Traceback:\n%s", traceback.format_exc())
        logger.error("===== CHECK IN END (ATTRIBUTE ERROR) =====")
        return jsonify(
            success=False,
            message="System configuration error. Please contact support."
        ), 500
    except Exception as exc:
        logger.error("===== CHECK IN EXCEPTION =====")
        logger.error("Exception Type: %s", type(exc).__name__)
        logger.error("Exception Message: %s", str(exc))
        import traceback
        logger.error("Full Traceback:\n%s", traceback.format_exc())
        # Log exception details for debugging
        import sys
        exc_info = sys.exc_info()
        logger.error("Exception Details: %s", exc_info)
        logger.error("===== CHECK IN END (EXCEPTION) =====")
        return jsonify(
            success=False,
            message=f"Check-in failed: System error occurred. [{type(exc).__name__}: {str(exc)}]"
        ), 500


# ── Check Out (AJAX) ─────────────────────────────────────────────────

@attendance_bp.route("/checkout", methods=["POST"])
@login_required
def checkout():
    """
    Check-out endpoint — validates GPS + photo, updates attendance record.
    """
    logger.info("===== CHECK OUT START =====")
    logger.info("User ID: %s", current_user.id)
    
    try:
        employee = _emp_repo.get_by_user_id(current_user.id)
        if not employee:
            logger.error("CHECK OUT FAILED: Employee profile not found")
            return jsonify(success=False, message="Employee profile not found."), 400
        
        logger.info("Employee ID: %s", employee.id)
        logger.info("Employee Name: %s", employee.full_name)

        # DEVICE VALIDATION — check if request is from allowed office computer
        if _device_svc.is_device_check_enabled():
            is_allowed, device_msg = _device_svc.is_allowed_device()
            logger.info("Device validation: allowed=%s, ip=%s", is_allowed, _device_svc.get_client_ip())
            if not is_allowed:
                logger.error("CHECK OUT FAILED: %s", device_msg)
                return jsonify(success=False, message=device_msg), 403  # 403 Forbidden

        # MANDATORY PHOTO VALIDATION — check if checkout photo was uploaded
        from app.models.attendance_photo import AttendancePhoto  # noqa: PLC0415
        today = date.today()
        
        # Get today's attendance
        attendance_today = _repo.get_today(employee.id, today)
        logger.info("Existing attendance today: %s", attendance_today)
        
        if not attendance_today:
            logger.error("CHECK OUT FAILED: No attendance record for today")
            return jsonify(
                success=False,
                message="No check-in found for today. Please check in first."
            ), 400
        
        if not attendance_today.check_in_time:
            logger.error("CHECK OUT FAILED: No check-in time")
            return jsonify(
                success=False,
                message="No check-in found for today. Please check in first."
            ), 400
        
        # Check if checkout photo exists
        photo = AttendancePhoto.query.filter_by(
            attendance_id=attendance_today.id
        ).first()
        logger.info("Photo record found: %s", photo)
        
        # Use getattr for safety - checkout_image_data might not exist on old Render DBs
        has_checkout_photo = bool(photo and getattr(photo, 'checkout_image_data', None))
        
        if not has_checkout_photo:
            logger.error("CHECK OUT FAILED: No checkout photo found")
            return jsonify(
                success=False,
                message="⚠️ Proof Photo is required to mark attendance. Please upload your GPS Map Camera selfie for check-out first."
            ), 400
        
        logger.info("Checkout photo validation PASSED")

        lat = request.form.get("latitude", "")
        lon = request.form.get("longitude", "")
        acc = request.form.get("accuracy", "")
        
        logger.info("GPS Data - Lat: %s, Lon: %s, Accuracy: %s", lat, lon, acc)

        ok, message, attendance, gps_detail = _svc.check_out(employee, lat, lon, acc)
        
        logger.info("Service check_out result: ok=%s, message=%s", ok, message)
        
        if ok:
            logger.info("CHECK OUT SUCCESS: attendance_id=%s, time=%s", 
                       attendance.id, attendance.check_out_time)
            logger.info("===== CHECK OUT END (SUCCESS) =====")
            
            # Ensure check_out_time is a datetime object before formatting
            checkout_time_str = "—"
            if attendance.check_out_time:
                try:
                    if hasattr(attendance.check_out_time, 'strftime'):
                        checkout_time_str = attendance.check_out_time.strftime("%H:%M")
                    else:
                        # Fallback if somehow it's a string
                        checkout_time_str = str(attendance.check_out_time)[-5:] if isinstance(attendance.check_out_time, str) else "—"
                except Exception as fmt_err:
                    logger.error("TIME FORMATTING ERROR: %s", str(fmt_err))
                    checkout_time_str = "—"
            
            return jsonify(
                success=True,
                message=message,
                time=checkout_time_str,
                working=attendance.working_hours_display,
                overtime_minutes=attendance.overtime_minutes or 0,
                gps=gps_detail,
            )
        
        logger.error("CHECK OUT FAILED: %s", message)
        logger.info("===== CHECK OUT END (FAILED) =====")
        return jsonify(success=False, message=message, gps=gps_detail), 400
        
    except Exception as exc:
        logger.error("===== CHECK OUT EXCEPTION =====")
        logger.error("Exception Type: %s", type(exc).__name__)
        logger.error("Exception Message: %s", str(exc))
        import traceback
        logger.error("Traceback:\n%s", traceback.format_exc())
        logger.error("===== CHECK OUT END (EXCEPTION) =====")
        return jsonify(
            success=False,
            message=f"Check-out failed: {str(exc)}",
            error_type=type(exc).__name__
        ), 500


# ── Photo Upload (AJAX) ──────────────────────────────────────────────

@attendance_bp.route("/upload-photo", methods=["POST"])
@login_required
def upload_photo():
    """
    Upload check-in proof photo.
    Returns updated state so frontend can sync button immediately.
    ALWAYS returns JSON, never HTML error pages.
    """
    logger.info("===== PHOTO UPLOAD START =====")
    
    try:
        employee = _emp_repo.get_by_user_id(current_user.id)
        if not employee:
            logger.error("UPLOAD FAILED: Employee not found")
            return jsonify(success=False, message="Employee profile not found."), 400
        
        logger.info("Employee ID: %s", employee.id)

        file = request.files.get("photo")
        if not file:
            logger.error("UPLOAD FAILED: No file in request")
            return jsonify(success=False, message="No file received."), 400
        
        logger.info("File received: %s", file.filename)
        file.seek(0)  # Reset file pointer

        try:
            ok, message, photo_url = _svc.upload_photo(employee, file)
        except Exception as svc_exc:
            logger.error("Service exception: %s", str(svc_exc))
            logger.error("Traceback: %s", traceback.format_exc())
            return jsonify(success=False, message="Upload service error: " + str(svc_exc)), 500
        
        if ok:
            # Get updated status to return to frontend
            from app.models.attendance_photo import AttendancePhoto  # noqa: PLC0415
            today = date.today()
            attendance_today = _repo.get_today(employee.id, today)
            
            has_photo = False
            if attendance_today and attendance_today.id:
                photo_rec = AttendancePhoto.query.filter_by(
                    attendance_id=attendance_today.id
                ).first()
                has_photo = bool(photo_rec and photo_rec.image_data)
            
            logger.info("UPLOAD SUCCESS: has_photo=%s", has_photo)
            logger.info("===== PHOTO UPLOAD END (SUCCESS) =====")
            
            return jsonify(
                success=True,
                message=message,
                photo_url=photo_url,
                has_photo=has_photo,
                can_check_in=bool(attendance_today and not attendance_today.check_in_time)
            )
        
        logger.error("UPLOAD FAILED: %s", message)
        logger.info("===== PHOTO UPLOAD END (FAILED) =====")
        return jsonify(success=False, message=message), 400
        
    except Exception as exc:
        logger.error("===== PHOTO UPLOAD TOP-LEVEL EXCEPTION =====")
        logger.error("Exception Type: %s", type(exc).__name__)
        logger.error("Exception Message: %s", str(exc))
        logger.error("Traceback:\n%s", traceback.format_exc())
        logger.error("===== PHOTO UPLOAD END (EXCEPTION) =====")
        
        return jsonify(
            success=False,
            message="Upload failed: " + str(exc)
        ), 500


# ── Upload Checkout Photo (AJAX) ─────────────────────────────────────

@attendance_bp.route("/upload-checkout-photo", methods=["POST"])
@login_required
def upload_checkout_photo():
    """
    Upload check-out proof photo.
    Returns updated state so frontend can sync button immediately.
    ALWAYS returns JSON, never HTML error pages.
    WORKS EXACTLY LIKE CHECK-IN: photo uploaded first, then checkout time recorded.
    """
    logger.info("===== CHECKOUT PHOTO UPLOAD START =====")
    
    try:
        employee = _emp_repo.get_by_user_id(current_user.id)
        if not employee:
            logger.error("CHECKOUT UPLOAD FAILED: Employee not found")
            return jsonify(success=False, message="Employee profile not found."), 400
        
        logger.info("Employee ID: %s", employee.id)

        file = request.files.get("photo")
        if not file:
            logger.error("CHECKOUT UPLOAD FAILED: No file in request")
            return jsonify(success=False, message="No file received."), 400
        
        logger.info("Checkout file received: %s", file.filename)
        file.seek(0)  # Reset file pointer

        try:
            ok, message, photo = _svc.upload_checkout_photo(employee, file)
        except Exception as svc_exc:
            logger.error("Service exception: %s", str(svc_exc))
            logger.error("Traceback: %s", traceback.format_exc())
            return jsonify(success=False, message="Checkout upload service error: " + str(svc_exc)), 500
        
        if ok:
            # Get updated state to return to frontend
            from app.models.attendance_photo import AttendancePhoto  # noqa: PLC0415
            today = date.today()
            attendance_today = _repo.get_today(employee.id, today)
            
            has_checkout_photo = False
            if attendance_today and attendance_today.id:
                photo_rec = AttendancePhoto.query.filter_by(
                    attendance_id=attendance_today.id
                ).first()
                has_checkout_photo = bool(photo_rec and photo_rec.checkout_image_data)
            
            logger.info("CHECKOUT UPLOAD SUCCESS: has_checkout_photo=%s", has_checkout_photo)
            logger.info("===== CHECKOUT PHOTO UPLOAD END (SUCCESS) =====")
            
            return jsonify(
                success=True,
                message=message,
                has_checkout_photo=has_checkout_photo,
                can_check_out=bool(
                    attendance_today 
                    and attendance_today.check_in_time 
                    and not attendance_today.check_out_time
                )
            )
        
        logger.error("CHECKOUT UPLOAD FAILED: %s", message)
        logger.info("===== CHECKOUT PHOTO UPLOAD END (FAILED) =====")
        return jsonify(success=False, message=message), 400
        
    except Exception as exc:
        logger.error("===== CHECKOUT PHOTO UPLOAD TOP-LEVEL EXCEPTION =====")
        logger.error("Exception Type: %s", type(exc).__name__)
        logger.error("Exception Message: %s", str(exc))
        logger.error("Traceback:\n%s", traceback.format_exc())
        logger.error("===== CHECKOUT PHOTO UPLOAD END (EXCEPTION) =====")
        
        return jsonify(
            success=False,
            message="Checkout upload failed: " + str(exc)
        ), 500


# ── Capture Selfie (Live Camera) ────────────────────────────────────

@attendance_bp.route("/capture-selfie", methods=["POST"])
@login_required
def capture_selfie():
    """
    Receive base64 selfie from live camera capture.
    Store in AttendancePhoto.image_data (check-in) or checkout_image_data (check-out).
    
    Expected JSON:
    {
        "selfie": "data:image/jpeg;base64,...",
        "type": "checkin" | "checkout"
    }
    
    Returns: {success, photo_id, has_photo, has_checkout_photo}
    
    ALWAYS returns JSON, never HTML errors.
    """
    from app.models.attendance_photo import AttendancePhoto  # noqa: PLC0415
    from app.extensions.database import db  # noqa: PLC0415

    logger.info("===== CAPTURE SELFIE START =====")
    logger.info("User ID: %s", current_user.id)
    
    try:
        # Ensure we always return JSON, never HTML
        response_headers = {'Content-Type': 'application/json'}
        
        # Check authentication first
        if not current_user or not current_user.is_authenticated:
            logger.error("CAPTURE SELFIE FAILED: User not authenticated")
            return jsonify(success=False, message="Authentication required. Please log in again."), 401
        
        employee = _emp_repo.get_by_user_id(current_user.id)
        if not employee:
            logger.error("CAPTURE SELFIE FAILED: Employee not found")
            return jsonify(success=False, message="Employee profile not found."), 400
        
        logger.info("Employee ID: %s", employee.id)
        
        # Get JSON payload with safe parsing
        try:
            data = request.get_json(silent=True) or {}
        except Exception as json_err:
            logger.error("CAPTURE SELFIE FAILED: JSON parsing error: %s", str(json_err))
            return jsonify(success=False, message="Invalid JSON format."), 400
        
        selfie_base64 = data.get("selfie", "").strip()
        capture_type = data.get("type", "checkin").lower()
        
        if not selfie_base64:
            logger.error("CAPTURE SELFIE FAILED: No selfie data")
            return jsonify(success=False, message="No selfie data received."), 400
        
        if not selfie_base64.startswith("data:image"):
            logger.error("CAPTURE SELFIE FAILED: Invalid data URL format")
            return jsonify(success=False, message="Invalid image data format."), 400
        
        if capture_type not in ("checkin", "checkout"):
            logger.error("CAPTURE SELFIE FAILED: Invalid type: %s", capture_type)
            return jsonify(success=False, message="Invalid capture type."), 400
        
        logger.info("Capture type: %s", capture_type)
        logger.info("Selfie data length: %d bytes", len(selfie_base64))
        
        # Get or create today's attendance
        today = date.today()
        attendance_today = _repo.get_today(employee.id, today)
        
        if not attendance_today:
            logger.info("Creating new attendance record for today")
            from app.models.attendance import Attendance  # noqa: PLC0415

            attendance_today = Attendance(
                employee_id=employee.id,
                date=today,
                status="pending",
            )
            db.session.add(attendance_today)
            db.session.flush()
            logger.info("Attendance created: %s", attendance_today.id)
        
        # Get or create photo record
        photo = AttendancePhoto.query.filter_by(
            attendance_id=attendance_today.id
        ).first()
        
        if not photo:
            logger.info("Creating new photo record")
            photo = AttendancePhoto(
                attendance_id=attendance_today.id,
                employee_id=employee.id,
                file_path="",
            )
            db.session.add(photo)
            db.session.flush()  # get ID without committing yet
            logger.info("Photo record created (pending commit): %s", photo.id)
        
        # Store selfie in appropriate field — always overwrite (retake support)
        if capture_type == "checkin":
            photo.image_data = selfie_base64
            logger.info("Stored check-in selfie")
        else:  # checkout
            photo.checkout_image_data = selfie_base64
            logger.info("Stored check-out selfie")
        
        db.session.commit()
        logger.info("Selfie saved successfully")
        
        # Return state for frontend
        has_photo = bool(photo.image_data)
        has_checkout_photo = bool(photo.checkout_image_data)
        
        logger.info("CAPTURE SELFIE SUCCESS: photo_id=%s, has_photo=%s, has_checkout_photo=%s",
                   photo.id, has_photo, has_checkout_photo)
        logger.info("===== CAPTURE SELFIE END (SUCCESS) =====")
        
        return jsonify(
            success=True,
            photo_id=photo.id,
            has_photo=has_photo,
            has_checkout_photo=has_checkout_photo,
            message="Selfie captured successfully."
        )
    
    except Exception as exc:
        logger.error("===== CAPTURE SELFIE EXCEPTION =====")
        logger.error("Exception Type: %s", type(exc).__name__)
        logger.error("Exception Message: %s", str(exc))
        logger.error("Traceback:\n%s", traceback.format_exc())
        logger.error("===== CAPTURE SELFIE END (EXCEPTION) =====")
        
        try:
            db.session.rollback()
        except Exception as rollback_err:
            logger.error("Rollback failed: %s", str(rollback_err))
        
        return jsonify(
            success=False,
            message=f"Selfie capture failed: {str(exc)}"
        ), 500


# ── Generate Proof Image ─────────────────────────────────────────────

@attendance_bp.route("/generate-proof-image", methods=["POST"])
@login_required
def generate_proof_image():
    """
    Generate professional attendance proof image using captured selfie + GPS data.
    
    Expected JSON:
    {
        "type": "checkin" | "checkout",
        "latitude": 18.5204,
        "longitude": 73.8567,
        "accuracy": 25.0,
        "distance_metres": 15.5
    }
    
    Returns: {success, proof_image_url, proof_size_bytes}
    """
    from app.models.attendance_photo import AttendancePhoto  # noqa: PLC0415
    from .proof_image_generator import ProofImageGenerator  # noqa: PLC0415
    
    logger.info("===== GENERATE PROOF IMAGE START =====")
    logger.info("User ID: %s", current_user.id)
    
    try:
        employee = _emp_repo.get_by_user_id(current_user.id)
        if not employee:
            logger.error("PROOF IMAGE FAILED: Employee not found")
            return jsonify(success=False, message="Employee profile not found."), 400
        
        logger.info("Employee ID: %s, Name: %s, Code: %s", 
                   employee.id, employee.full_name, employee.employee_code)
        
        # Parse request
        data = request.get_json() or {}
        proof_type = data.get("type", "checkin").lower()
        latitude = data.get("latitude")
        longitude = data.get("longitude")
        accuracy = data.get("accuracy")
        distance_metres = data.get("distance_metres")
        
        if not all([latitude, longitude, accuracy is not None, distance_metres is not None]):
            logger.error("PROOF IMAGE FAILED: Missing GPS data")
            return jsonify(success=False, message="Missing GPS data."), 400
        
        logger.info("Proof type: %s, GPS: (%.6f, %.6f), Accuracy: %.1f, Distance: %.1f",
                   proof_type, latitude, longitude, accuracy, distance_metres)
        
        # Get today's attendance and photo
        today = date.today()
        attendance_today = _repo.get_today(employee.id, today)
        
        if not attendance_today:
            logger.error("PROOF IMAGE FAILED: No attendance record for today")
            return jsonify(success=False, message="No attendance record found."), 400
        
        photo = AttendancePhoto.query.filter_by(
            attendance_id=attendance_today.id
        ).first()
        
        if not photo:
            logger.error("PROOF IMAGE FAILED: No photo record found")
            return jsonify(success=False, message="No photo record found."), 400
        
        # Get the appropriate selfie based on proof_type
        if proof_type == "checkin":
            selfie_base64 = photo.image_data
            if not selfie_base64:
                logger.error("PROOF IMAGE FAILED: No check-in selfie found")
                return jsonify(success=False, message="No check-in selfie found."), 400
            logger.info("Using check-in selfie")
        else:  # checkout
            selfie_base64 = photo.checkout_image_data
            if not selfie_base64:
                logger.error("PROOF IMAGE FAILED: No check-out selfie found")
                return jsonify(success=False, message="No check-out selfie found."), 400
            logger.info("Using check-out selfie")
        
        # Get office info
        office = _repo.get_office_for_employee(employee)
        if not office:
            logger.warning("PROOF IMAGE: No office settings found, using defaults")
            office_name = "Head Office"
            office_address = "Office Address"
            allowed_radius = 100
        else:
            # Handle both dict and object types
            if isinstance(office, dict):
                office_name = office.get('name', 'Head Office')
                office_address = office.get('address', 'Office Address')
                allowed_radius = int(office.get('radius_metres', 100) or 100)
            else:
                office_name = getattr(office, 'name', None) or "Head Office"
                office_address = getattr(office, 'address', None) or "Office Address"
                allowed_radius = int(getattr(office, 'radius_metres', 100) or 100)
            logger.info("Office: %s, Radius: %d", office_name, allowed_radius)
        
        # Get device info from user agent
        device_info = request.headers.get("User-Agent", "Unknown Device")
        logger.info("Device info: %s", device_info)
        
        # Generate proof image
        logger.info("Generating proof image...")
        generator = ProofImageGenerator()
        
        proof_image_base64 = generator.generate(
            selfie_base64=selfie_base64,
            employee_name=employee.full_name or "Unknown",
            employee_code=employee.employee_code or "N/A",
            department=employee.department or "N/A",
            designation=employee.designation or "N/A",
            office_name=office_name,
            office_address=office_address,
            check_type=proof_type,
            latitude=float(latitude),
            longitude=float(longitude),
            accuracy=float(accuracy),
            distance_metres=float(distance_metres),
            allowed_radius=int(allowed_radius),
            device_info=device_info,
            attendance_id=str(attendance_today.id),
        )
        
        logger.info("Proof image generated, size: %d bytes", len(proof_image_base64))
        
        # Store proof image back to photo record (overwrites selfie)
        if proof_type == "checkin":
            photo.image_data = proof_image_base64
            logger.info("Stored check-in proof image")
        else:  # checkout
            photo.checkout_image_data = proof_image_base64
            logger.info("Stored check-out proof image")
        
        from app.extensions.database import db  # noqa: PLC0415
        db.session.commit()
        logger.info("Proof image saved to database")
        
        logger.info("GENERATE PROOF IMAGE SUCCESS: proof_image_url (data URI)")
        logger.info("===== GENERATE PROOF IMAGE END (SUCCESS) =====")
        
        return jsonify(
            success=True,
            proof_image_url=proof_image_base64,
            proof_size_bytes=len(proof_image_base64),
            message="Proof image generated successfully."
        )
    
    except Exception as exc:
        logger.error("===== GENERATE PROOF IMAGE EXCEPTION =====")
        logger.error("Exception Type: %s", type(exc).__name__)
        logger.error("Exception Message: %s", str(exc))
        logger.error("Traceback:\n%s", traceback.format_exc())
        logger.error("===== GENERATE PROOF IMAGE END (EXCEPTION) =====")
        
        return jsonify(
            success=False,
            message=f"Proof image generation failed: {str(exc)}"
        ), 500


# ── Attendance history ────────────────────────────────────────────────

@attendance_bp.route("/history")
@login_required
def history():
    employee = _emp_repo.get_by_user_id(current_user.id)
    if not employee:
        flash("Employee profile not found.", "warning")
        return redirect(url_for("dashboard.index"))

    page       = request.args.get("page",       1,  type=int)
    start_date = request.args.get("start_date", "").strip()
    end_date   = request.args.get("end_date",   "").strip()
    status     = request.args.get("status",     "").strip()

    # Parse date strings
    from datetime import datetime as _dt
    parsed_start = None
    parsed_end   = None
    try:
        if start_date:
            parsed_start = _dt.strptime(start_date, "%Y-%m-%d").date()
        if end_date:
            parsed_end = _dt.strptime(end_date, "%Y-%m-%d").date()
    except ValueError:
        parsed_start = parsed_end = None

    pagination = _repo.get_history_with_photos(
        employee.id,
        start_date=parsed_start,
        end_date=parsed_end,
        status=status,
        page=page,
        per_page=30,
    )

    # Pass office settings so the template can show shift timing
    office = _repo.get_office_for_employee(employee)

    return render_template(
        "attendance/history.html",
        title="Attendance History",
        employee=employee,
        pagination=pagination,
        start_date=start_date,
        end_date=end_date,
        status=status,
        office=office,
    )


# ── Export Attendance History to Excel ───────────────────────────────

@attendance_bp.route("/history/export")
@login_required
def export_history():
    """
    Export attendance history to Excel (.xlsx).
    Respects the same filters as the history page.
    Uses openpyxl — no temp file written to disk.
    """
    import io
    from datetime import datetime as _dt
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    from flask import send_file

    employee = _emp_repo.get_by_user_id(current_user.id)
    if not employee:
        return jsonify(success=False, message="Employee profile not found."), 400

    # Parse filters (same as history route)
    start_date = request.args.get("start_date", "").strip()
    end_date   = request.args.get("end_date",   "").strip()
    status     = request.args.get("status",     "").strip()

    parsed_start = None
    parsed_end   = None
    try:
        if start_date:
            parsed_start = _dt.strptime(start_date, "%Y-%m-%d").date()
        if end_date:
            parsed_end = _dt.strptime(end_date, "%Y-%m-%d").date()
    except ValueError:
        pass

    # Fetch all matching records (no pagination)
    records = _repo.get_history_all_filtered(
        employee.id,
        start_date=parsed_start,
        end_date=parsed_end,
        status=status,
    )

    if not records:
        return jsonify(
            success=False,
            message="No attendance records available for export."
        ), 404

    # ── Build Excel workbook ──────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance History"

    # Styles
    header_font    = Font(bold=True, color="FFFFFF", size=11)
    header_fill    = PatternFill(start_color="1A3C6E", end_color="1A3C6E", fill_type="solid")
    header_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_align   = Alignment(horizontal="center", vertical="center")
    thin_border    = Border(
        left=Side(style="thin", color="DEE2E6"),
        right=Side(style="thin", color="DEE2E6"),
        top=Side(style="thin", color="DEE2E6"),
        bottom=Side(style="thin", color="DEE2E6"),
    )
    alt_fill = PatternFill(start_color="F4F6F9", end_color="F4F6F9", fill_type="solid")

    # ── Headers ───────────────────────────────────────────────────────
    headers = [
        ("Employee ID",      14),
        ("Employee Name",    22),
        ("Date",             14),
        ("Day",              10),
        ("Check In (IST)",   16),
        ("Check Out (IST)",  16),
        ("Working Hours",    14),
        ("Overtime",         12),
        ("Status",           14),
        ("Late",             10),
        ("Late By (min)",    13),
        ("Remarks",          20),
    ]

    for col_idx, (header_text, col_width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font    = header_font
        cell.fill    = header_fill
        cell.alignment = header_align
        cell.border  = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[1].height = 28

    # Freeze header row
    ws.freeze_panes = "A2"

    # ── Data rows ─────────────────────────────────────────────────────
    def to_ist(dt):
        """Convert naive UTC datetime to IST string HH:MM AM/PM."""
        if not dt:
            return "—"
        # IST = UTC + 5:30
        from datetime import timedelta
        ist = dt + timedelta(hours=5, minutes=30)
        return ist.strftime("%I:%M %p")

    def fmt_date(d):
        return d.strftime("%d-%m-%Y") if d else "—"

    def fmt_hours(minutes):
        if not minutes:
            return "—"
        h, m = divmod(minutes, 60)
        return f"{h}h {m:02d}m"

    def fmt_overtime(minutes):
        if not minutes:
            return "—"
        h, m = divmod(minutes, 60)
        return f"+{h}h {m:02d}m" if h else f"+{m}m"

    def fmt_status(status_str):
        return status_str.replace("_", " ").title() if status_str else "—"

    for row_idx, att in enumerate(records, start=2):
        is_alt = (row_idx % 2 == 0)
        row_fill = alt_fill if is_alt else None

        row_data = [
            att.employee.employee_code,
            att.employee.full_name,
            fmt_date(att.date),
            att.date.strftime("%A") if att.date else "—",
            to_ist(att.check_in_time),
            to_ist(att.check_out_time),
            fmt_hours(att.working_minutes),
            fmt_overtime(att.overtime_minutes),
            fmt_status(att.status),
            "Yes" if att.is_late else "No",
            att.late_minutes if att.is_late else 0,
            "Half Day" if att.is_half_day else (
                "Early Leave" if att.is_early_leave else ""
            ),
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = thin_border
            cell.alignment = center_align if col_idx not in (2,) else Alignment(vertical="center")
            if row_fill:
                cell.fill = row_fill

        # Highlight late rows
        if att.is_late:
            ws.cell(row=row_idx, column=10).font = Font(color="D97706", bold=True)

        # Colour status cell
        status_colors = {
            "present":   "198754",
            "absent":    "DC3545",
            "half_day":  "FFC107",
            "on_leave":  "0DCAF0",
            "holiday":   "6C757D",
            "weekend":   "ADB5BD",
        }
        color = status_colors.get(att.status, "212529")
        ws.cell(row=row_idx, column=9).font = Font(color=color, bold=True)

    # ── Summary row ───────────────────────────────────────────────────
    total_row = len(records) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, color="1A3C6E")
    ws.cell(row=total_row, column=1).fill = PatternFill(
        start_color="EFF6FF", end_color="EFF6FF", fill_type="solid"
    )
    ws.cell(row=total_row, column=7,
            value=fmt_hours(sum(r.working_minutes or 0 for r in records))
    ).font = Font(bold=True, color="198754")

    # ── Save to in-memory buffer ──────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"Attendance_History_{_dt.now().strftime('%Y-%m-%d')}.xlsx"

    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# ── Office Settings (HR / Admin only) ────────────────────────────────

@attendance_bp.route("/settings", methods=["GET", "POST"])
@login_required
@hr_required
def settings():
    office = _office_svc.get_or_create_default()
    form   = OfficeSettingsForm(obj=office)

    if form.validate_on_submit():
        ok, msg = _office_svc.update(
            office_id=office.id,
            data={
                "name":                       form.name.data,
                "address":                    form.address.data,
                "latitude":                   form.latitude.data,
                "longitude":                  form.longitude.data,
                "radius_metres":              form.radius_metres.data,
                "office_start_time":          form.office_start_time.data,
                "office_end_time":            form.office_end_time.data,
                "grace_period_minutes":       form.grace_period_minutes.data,
                "half_day_threshold_minutes": form.half_day_threshold_minutes.data,
                "overtime_threshold_minutes": form.overtime_threshold_minutes.data,
                "allow_remote_checkin":       form.allow_remote_checkin.data,
                "selfie_required":            form.selfie_required.data,
            },
            updated_by=current_user.id,
        )
        flash(msg, "success" if ok else "danger")
        if ok:
            return redirect(url_for("attendance.settings"))

    return render_template(
        "attendance/settings.html",
        title="Attendance Settings",
        form=form,
        office=office,
    )


# ── View proof photo (full page) ─────────────────────────────────────

@attendance_bp.route("/photo-view/<int:photo_id>")
@login_required
def view_photo(photo_id):
    """
    Display a proof photo as a full-page image view.
    Uses the base64 image_data stored in the DB — no filesystem needed.
    Falls back to serve_photo redirect for old file-based records.
    """
    from app.models.attendance_photo import AttendancePhoto  # noqa: PLC0415
    from flask import Response  # noqa: PLC0415

    photo = AttendancePhoto.query.filter_by(id=photo_id).first_or_404()

    # Security: only the employee who owns the photo can view it
    # (HR/Admin access handled separately via admin routes)
    employee = _emp_repo.get_by_user_id(current_user.id)
    if not employee or photo.employee_id != employee.id:
        # Allow if user is HR/Admin
        from app.constants.enums import UserRole  # noqa: PLC0415
        if not hasattr(current_user, 'role') or current_user.role not in (
            UserRole.ADMIN.value, UserRole.SUPER_ADMIN.value, UserRole.HR_MANAGER.value
        ):
            abort(403)

    if photo.image_data:
        # Serve a minimal HTML page that displays the image full-screen
        date_str = ""
        if photo.uploaded_at:
            import pytz  # noqa: PLC0415
            ist = photo.uploaded_at.astimezone(pytz.timezone("Asia/Kolkata"))
            date_str = ist.strftime("%d %b %Y %I:%M %p IST")

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>Proof Photo — Smart HRMS</title>
  <style>
    *{{margin:0;padding:0;box-sizing:border-box}}
    body{{background:#1a1a2e;display:flex;flex-direction:column;align-items:center;
         justify-content:center;min-height:100vh;font-family:system-ui,sans-serif}}
    .info{{color:#94a3b8;font-size:.85rem;margin-bottom:16px;text-align:center}}
    .info strong{{color:#e2e8f0;font-size:1rem;display:block;margin-bottom:4px}}
    img{{max-width:95vw;max-height:85vh;object-fit:contain;border-radius:12px;
         box-shadow:0 8px 40px rgba(0,0,0,.6)}}
    .back{{margin-top:20px;color:#60a5fa;text-decoration:none;font-size:.85rem;
           padding:8px 20px;border:1px solid #3b82f6;border-radius:8px}}
    .back:hover{{background:#1d4ed8;color:#fff}}
  </style>
</head>
<body>
  <div class="info">
    <strong>Proof Photo</strong>
    {date_str}
  </div>
  <img src="{photo.image_data}" alt="Proof photo">
  <a href="javascript:history.back()" class="back">← Back</a>
</body>
</html>"""
        return Response(html, mimetype="text/html")

    # Old file-based record — redirect to serve_photo
    if photo.file_path:
        return redirect(url_for("attendance.serve_photo", filename=photo.file_path))

    abort(404)


# ── Serve attendance proof photos ────────────────────────────────────

@attendance_bp.route("/photo/<path:filename>")
@login_required
def serve_photo(filename):
    """
    Serve an attendance proof photo from the uploads folder.
    Only authenticated users can access this endpoint.
    Security: prevent directory traversal via normpath check.
    """
    from flask import current_app  # noqa: PLC0415
    from pathlib import Path  # noqa: PLC0415
    import logging as _log  # noqa: PLC0415

    _logger = _log.getLogger("attendance")

    # Security: prevent directory traversal
    safe_name = os.path.normpath(filename)
    if ".." in safe_name or safe_name.startswith("/"):
        abort(403)

    # Validate extension
    ext = os.path.splitext(safe_name)[1].lower().lstrip(".")
    if ext not in {"jpg", "jpeg", "png", "webp", "gif"}:
        abort(403)

    # Resolve UPLOAD_FOLDER to absolute path — critical for Render deployment
    # where relative paths resolve from an unpredictable CWD.
    raw_folder = current_app.config.get("UPLOAD_FOLDER", "./instance/uploads")
    upload_folder = Path(raw_folder)
    if not upload_folder.is_absolute():
        clean = raw_folder
        if clean.startswith("./"):
            clean = clean[2:]
        upload_folder = Path(current_app.root_path).parent / clean

    upload_folder = upload_folder.resolve()
    target = (upload_folder / safe_name).resolve()

    # Ensure resolved path is still inside upload_folder (traversal guard)
    try:
        target.relative_to(upload_folder)
    except ValueError:
        abort(403)

    # Debug log — helps diagnose 404s on Render
    _logger.info(
        "SERVE_PHOTO | file=%s | upload_folder=%s | target=%s | exists=%s",
        safe_name, upload_folder, target, target.exists()
    )

    if not target.exists():
        _logger.warning("PHOTO_NOT_FOUND | path=%s", target)
        abort(404)

    try:
        return send_from_directory(str(upload_folder), safe_name)
    except Exception as exc:
        _logger.error("PHOTO_SERVE_ERROR | %s", exc)
        abort(404)
