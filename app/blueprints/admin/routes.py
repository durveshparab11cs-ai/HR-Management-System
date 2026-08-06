"""
blueprints/admin/routes.py
============================
Admin panel routes — dashboard, office settings, user management.
"""

from datetime import date
from flask import flash, jsonify, redirect, render_template, request, url_for
from flask_login import current_user, login_required

from app.core.security import admin_required, roles_required
from app.constants.enums import UserRole
from app.blueprints.attendance.repository import AttendanceRepository
from app.blueprints.employees.repository import EmployeeRepository
from app.blueprints.leave.repository import LeaveRepository
from .forms import OfficeSettingsForm
from .service import AdminService
from . import admin_bp

# TEMPORARY DEBUGGING ENDPOINT - Remove after role update confirmed
@admin_bp.route("/debug-users", methods=["GET"])
def debug_users():
    """Temporary endpoint to check user roles in production database."""
    from app.models.user import User
    from app.extensions.database import db
    
    try:
        user1 = User.query.filter_by(username='e_2512012').first()
        user2 = User.query.filter_by(username='e_2603025').first()
        
        return {
            "e_2512012": {
                "exists": user1 is not None,
                "role": user1.role if user1 else None,
                "email": user1.email if user1 else None
            },
            "e_2603025": {
                "exists": user2 is not None,
                "role": user2.role if user2 else None,
                "email": user2.email if user2 else None
            }
        }, 200
    except Exception as e:
        return {"error": str(e)}, 500

_att   = AttendanceRepository()
_emp   = EmployeeRepository()
_leave = LeaveRepository()
_svc   = AdminService()


@admin_bp.route("/")
@admin_bp.route("")
@login_required
@admin_required
def index():
    today = date.today()
    from datetime import datetime
    now = datetime.now()
    try:
        total_employees   = _emp.count_total()
    except Exception:
        total_employees   = 0
    try:
        checked_in_today  = _att.count_checked_in_today(today)
    except Exception:
        checked_in_today  = 0
    try:
        checked_out_today = _att.count_checked_out_today(today)
    except Exception:
        checked_out_today = 0
    try:
        late_today        = _att.count_late_today(today)
    except Exception:
        late_today        = 0
    absent_today = max(0, total_employees - checked_in_today)
    try:
        pending_leaves    = _leave.count_pending()
    except Exception:
        pending_leaves    = 0
    try:
        pending_halfdays  = _leave.count_pending_halfdays()
    except Exception:
        pending_halfdays  = 0
    try:
        pending_early     = _leave.count_pending_earlyleaves()
    except Exception:
        pending_early     = 0
    try:
        today_records     = _att.get_all_today(today)
    except Exception:
        today_records     = []
    try:
        recent_requests   = _leave.get_pending(page=1, per_page=5).items
    except Exception:
        recent_requests   = []

    return render_template(
        "admin/index.html",
        title="Admin Dashboard",
        today=today,
        now=now,
        total_employees=total_employees,
        checked_in_today=checked_in_today,
        checked_out_today=checked_out_today,
        late_today=late_today,
        absent_today=absent_today,
        pending_leaves=pending_leaves,
        pending_halfdays=pending_halfdays,
        pending_early=pending_early,
        today_records=today_records,
        recent_requests=recent_requests,
    )


@admin_bp.route("/attendance/all")
@login_required
@admin_required
def view_all_attendance():
    """
    View all employees' attendance with check-in and check-out proof photos.
    Shows today's attendance by default, with date filter and pagination.
    """
    from datetime import datetime as _dt
    from app.models.attendance import Attendance
    from app.models.attendance_photo import AttendancePhoto
    from app.models.employee import Employee
    from app.models.user import User
    from sqlalchemy import desc
    from app.extensions.database import db

    # Get date filter (default: today)
    date_str = request.args.get("date", "").strip()
    try:
        view_date = _dt.strptime(date_str, "%Y-%m-%d").date() if date_str else date.today()
    except ValueError:
        view_date = date.today()

    page = request.args.get("page", 1, type=int)
    per_page = 20

    # Query all attendance for the given date with employee data
    query = (
        db.session.query(Attendance, Employee, User, AttendancePhoto)
        .join(Employee, Attendance.employee_id == Employee.id)
        .join(User, Employee.user_id == User.id)
        .outerjoin(AttendancePhoto, Attendance.id == AttendancePhoto.attendance_id)
        .filter(
            Attendance.date == view_date,
            Attendance.is_deleted == False,
            Employee.is_deleted == False,
            User.is_deleted == False,
        )
        .order_by(desc(Attendance.check_in_time))
    )

    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    # Format attendance data for template
    attendance_data = []
    for att, emp, usr, photo in pagination.items:
        from datetime import timedelta

        # Convert times to IST
        def to_ist(dt):
            if not dt:
                return None
            return dt + timedelta(hours=5, minutes=30)

        check_in_ist = to_ist(att.check_in_time)
        check_out_ist = to_ist(att.check_out_time)

        attendance_data.append({
            'attendance': att,
            'employee': emp,
            'user': usr,
            'photo': photo,
            'check_in_time_ist': check_in_ist,
            'check_out_time_ist': check_out_ist,
            'working_hours_display': att.working_hours_display,
            'employee_code': emp.employee_code,
            'employee_name': usr.full_name,
            'department': emp.department or '—',
            'status_display': att.status.replace('_', ' ').title(),
        })

    return render_template(
        "admin/view_all_attendance.html",
        title="All Employee Attendance",
        view_date=view_date,
        date_str=view_date.strftime("%Y-%m-%d"),
        attendance_data=attendance_data,
        pagination=pagination,
        page=page,
    )


@admin_bp.route("/office-settings", methods=["GET", "POST"])
@login_required
@admin_required
def office_settings():
    office = _svc.get_or_create_default_office()
    form = OfficeSettingsForm(obj=office)
    if form.validate_on_submit():
        ok, msg = _svc.update_office_settings(office.id, {
            "name":                    form.name.data,
            "address":                 form.address.data,
            "latitude":                form.latitude.data,
            "longitude":               form.longitude.data,
            "radius_metres":           form.radius_metres.data,
            "office_start_time":       form.office_start_time.data,
            "office_end_time":         form.office_end_time.data,
            "grace_period_minutes":    form.grace_period_minutes.data,
            "half_day_threshold_minutes": form.half_day_threshold_minutes.data,
            "overtime_threshold_minutes": form.overtime_threshold_minutes.data,
            "allow_remote_checkin":    form.allow_remote_checkin.data,
            "selfie_required":         form.selfie_required.data,
        })
        flash(msg, "success" if ok else "danger")
        if ok:
            return redirect(url_for("admin.office_settings"))
    return render_template(
        "admin/office_settings.html",
        title="Office Settings",
        form=form,
        office=office,
    )


@admin_bp.route("/users")
@login_required
@admin_required
def users():
    try:
        page = request.args.get("page", 1, type=int)
        search = request.args.get("q", "")
        # For admin user management, we need to bypass department filtering
        # Query directly without department constraint
        from app.models.employee import Employee
        from app.models.user import User
        from sqlalchemy import or_
        
        query = Employee.query.filter(Employee.is_deleted == False)
        
        if search:
            term = f"%{search}%"
            query = (
                query.join(User, Employee.user_id == User.id, isouter=True)
                .filter(or_(
                    Employee.employee_code.ilike(term),
                    User.first_name.ilike(term),
                    User.last_name.ilike(term),
                    User.email.ilike(term),
                    Employee.mobile.ilike(term),
                ))
            )
        
        pagination = query.order_by(Employee.employee_code.asc()).paginate(
            page=page, per_page=25, error_out=False
        )
        
        return render_template(
            "admin/users.html",
            title="User Management",
            pagination=pagination,
            employees=pagination.items,
            search=search,
        )
    except Exception as e:
        import logging
        logging.error(f"USER_MANAGEMENT_ERROR: {str(e)}", exc_info=True)
        # Return empty list on error instead of crashing
        from flask import flash
        flash(f"Error loading users: {str(e)}", "warning")
        from app.extensions.database import db
        class EmptyPagination:
            items = []
            pages = 1
            page = 1
            per_page = 25
            total = 0
        return render_template(
            "admin/users.html",
            title="User Management",
            pagination=EmptyPagination(),
            employees=[],
            search="",
        )


@admin_bp.route("/audit-logs")
@login_required
@roles_required(UserRole.SUPER_ADMIN)
def audit_logs():
    return render_template("admin/audit_logs.html", title="Audit Logs")


@admin_bp.route("/leave-types")
@login_required
@admin_required
def leave_types():
    types = _leave.get_all_types()
    return render_template("admin/leave_types.html", title="Leave Types", leave_types=types)


# ── Employee Import Module ────────────────────────────────────────────

@admin_bp.route("/import-employees", methods=["GET", "POST"])
@login_required
@admin_required
def import_employees():
    """Upload Excel and import employee master data."""
    from .employee_import import EmployeeImportService
    from app.models.employee_master import EmployeeMaster

    import_svc = EmployeeImportService()
    result     = None
    total_masters = EmployeeMaster.query.count()
    registered    = EmployeeMaster.query.filter_by(is_registered=True).count()

    if request.method == "POST":
        file = request.files.get("excel_file")
        if not file or not file.filename:
            flash("Please select an Excel file to upload.", "warning")
        elif not file.filename.lower().endswith((".xlsx", ".xls")):
            flash("Only Excel files (.xlsx, .xls) are supported.", "danger")
        else:
            result = import_svc.import_from_file(file, imported_by=current_user.id)
            if result.get("success"):
                flash(result["message"], "success")
            else:
                flash(result["message"], "danger")
            total_masters = EmployeeMaster.query.count()
            registered    = EmployeeMaster.query.filter_by(is_registered=True).count()

    return render_template(
        "admin/import_employees.html",
        title="Import Employees",
        result=result,
        total_masters=total_masters,
        registered=registered,
    )


@admin_bp.route("/import-employees/preview", methods=["POST"])
@login_required
@admin_required
def preview_import():
    """AJAX: preview Excel without importing."""
    from .employee_import import EmployeeImportService
    import_svc = EmployeeImportService()
    file = request.files.get("excel_file")
    if not file or not file.filename:
        return jsonify(success=False, message="No file provided.")
    result = import_svc.preview(file)
    return jsonify(result)


@admin_bp.route("/employee-master")
@login_required
@admin_required
def employee_master():
    """View all employees in the master table."""
    from app.models.employee_master import EmployeeMaster
    from flask import request as _req
    
    search = _req.args.get("q", "").strip()
    status = _req.args.get("status", "")
    page   = _req.args.get("page", 1, type=int)

    q = EmployeeMaster.query
    if search:
        q = q.filter(
            EmployeeMaster.employee_code.ilike(f"%{search}%") |
            EmployeeMaster.employee_name.ilike(f"%{search}%") |
            EmployeeMaster.department.ilike(f"%{search}%")
        )
    if status == "registered":
        q = q.filter_by(is_registered=True)
    elif status == "pending":
        q = q.filter_by(is_registered=False)

    pagination = q.order_by(EmployeeMaster.employee_code).paginate(
        page=page, per_page=30, error_out=False
    )
    return render_template(
        "admin/employee_master.html",
        title="Employee Master",
        pagination=pagination,
        search=search,
        status=status,
    )


@admin_bp.route("/attendance/export")
@login_required
@admin_required
def export_daily_attendance():
    """
    Export full daily attendance for ALL employees to Excel.
    Query param: ?date=YYYY-MM-DD  (defaults to today)
    """
    import io
    from datetime import datetime as _dt
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import send_file
    from app.blueprints.attendance.repository import AttendanceRepository as AttRepo

    att_repo = AttRepo()

    date_str = request.args.get("date", "").strip()
    try:
        export_date = _dt.strptime(date_str, "%Y-%m-%d").date() if date_str else date.today()
    except ValueError:
        export_date = date.today()

    rows = att_repo.get_all_employees_attendance_for_date(export_date)

    if not rows:
        flash("No employee data available for export.", "warning")
        return redirect(url_for("admin.index"))

    # ── Build workbook ────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Attendance"

    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill(start_color="1A3C6E", end_color="1A3C6E", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left", vertical="center")
    thin         = Side(style="thin", color="DEE2E6")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill     = PatternFill(start_color="F4F6F9", end_color="F4F6F9", fill_type="solid")

    # ── Title row ─────────────────────────────────────────────────────
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = f"Daily Attendance Report — {export_date.strftime('%d %B %Y')}"
    title_cell.font  = Font(bold=True, size=13, color="1A3C6E")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Headers ───────────────────────────────────────────────────────
    headers = [
        ("#",              5),
        ("Emp Code",       12),
        ("Employee Name",  22),
        ("Department",     16),
        ("Designation",    16),
        ("Date",           12),
        ("Check In (IST)", 15),
        ("Check Out (IST)",15),
        ("Working Hours",  14),
        ("Status",         13),
        ("Late",           8),
        ("Late By (min)",  13),
    ]

    for col, (h, w) in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = border
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[2].height = 24
    ws.freeze_panes = "A3"

    # ── Helpers ───────────────────────────────────────────────────────
    def to_ist(dt):
        if not dt:
            return "—"
        from datetime import timedelta
        return (dt + timedelta(hours=5, minutes=30)).strftime("%I:%M %p")

    def fmt_hours(mins):
        if not mins:
            return "—"
        h, m = divmod(mins, 60)
        return f"{h}h {m:02d}m"

    status_colors = {
        "present":  "198754",
        "absent":   "DC3545",
        "half_day": "D97706",
        "on_leave": "0891B2",
        "holiday":  "6C757D",
        "weekend":  "ADB5BD",
    }

    # ── Data rows ─────────────────────────────────────────────────────
    summary = {"present": 0, "absent": 0, "late": 0, "on_leave": 0}

    for i, row in enumerate(rows, start=1):
        r   = i + 2
        att = row["attendance"]
        emp = row["employee"]
        usr = row["user"]
        is_alt = (i % 2 == 0)

        status_str = att.status if att else "absent"
        if status_str in summary:
            summary[status_str] += 1
        elif status_str not in summary:
            pass

        if att and att.check_in_time:
            summary["present"] = summary.get("present", 0)
        else:
            summary["absent"] = summary.get("absent", 0)

        if att and att.is_late:
            summary["late"] += 1

        row_data = [
            i,
            emp.employee_code,
            usr.full_name,
            emp.department or "—",
            emp.designation or "—",
            export_date.strftime("%d-%m-%Y"),
            to_ist(att.check_in_time) if att else "—",
            to_ist(att.check_out_time) if att else "—",
            fmt_hours(att.working_minutes) if att else "—",
            status_str.replace("_", " ").title(),
            "Yes" if (att and att.is_late) else "No",
            att.late_minutes if (att and att.is_late) else 0,
        ]

        for col, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border    = border
            cell.alignment = left_align if col in (3, 4, 5) else center_align
            if is_alt:
                cell.fill = alt_fill

        # Colour status
        st_cell = ws.cell(row=r, column=10)
        color   = status_colors.get(status_str, "212529")
        st_cell.font = Font(color=color, bold=True)

        # Highlight late
        if att and att.is_late:
            ws.cell(row=r, column=11).font = Font(color="D97706", bold=True)

    # ── Summary row ───────────────────────────────────────────────────
    sr = len(rows) + 3
    ws.cell(row=sr, column=1, value="SUMMARY").font = Font(bold=True, color="1A3C6E")
    summary_data = [
        ("Total Employees", len(rows), "1A3C6E"),
        ("Present",   sum(1 for r in rows if r["attendance"] and r["attendance"].check_in_time), "198754"),
        ("Absent",    sum(1 for r in rows if not r["attendance"] or not r["attendance"].check_in_time), "DC3545"),
        ("Late",      sum(1 for r in rows if r["attendance"] and r["attendance"].is_late), "D97706"),
        ("On Leave",  sum(1 for r in rows if r["attendance"] and r["attendance"].status == "on_leave"), "0891B2"),
    ]
    for col, (label, val, color) in enumerate(summary_data, start=1):
        lc = ws.cell(row=sr, column=col, value=label)
        lc.font  = Font(bold=True, color=color, size=9)
        lc.fill  = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
        vc = ws.cell(row=sr + 1, column=col, value=val)
        vc.font  = Font(bold=True, color=color, size=14)
        vc.fill  = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
        vc.alignment = Alignment(horizontal="center")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"Attendance_{export_date.strftime('%Y-%m-%d')}.xlsx"
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@admin_bp.route("/attendance/reset", methods=["POST"])
@login_required
@admin_required
def reset_attendance():
    """
    Reset ALL attendance data (development/testing only).
    Requires confirmation via JSON body: {"confirm": "DELETE ALL"}
    """
    import logging
    from app.models.attendance import Attendance
    from app.models.attendance_photo import AttendancePhoto
    from app.models.attendance_log import AttendanceLog
    from app.extensions.database import db

    logger = logging.getLogger("admin")

    data = request.get_json() or {}
    confirm = data.get("confirm", "")

    if confirm != "DELETE ALL":
        return jsonify(
            success=False,
            message="Confirmation failed. Send {\"confirm\": \"DELETE ALL\"} to proceed."
        ), 400

    try:
        # Count before deletion
        attendance_count = Attendance.query.count()
        photo_count = AttendancePhoto.query.count()
        log_count = AttendanceLog.query.count()

        logger.info(
            "ATTENDANCE_RESET_START | by_user=%s | att=%d | photos=%d | logs=%d",
            current_user.id, attendance_count, photo_count, log_count
        )

        # Delete in correct order (respect foreign keys)
        AttendanceLog.query.delete()
        AttendancePhoto.query.delete()
        Attendance.query.delete()
        db.session.commit()

        logger.info(
            "ATTENDANCE_RESET_SUCCESS | by_user=%s | deleted att=%d, photos=%d, logs=%d",
            current_user.id, attendance_count, photo_count, log_count
        )

        return jsonify(
            success=True,
            message=f"Successfully deleted {attendance_count} attendance records, {photo_count} photos, and {log_count} logs.",
            deleted={
                "attendance": attendance_count,
                "photos": photo_count,
                "logs": log_count
            }
        )

    except Exception as exc:
        db.session.rollback()
        logger.error("ATTENDANCE_RESET_FAILED | by_user=%s | error=%s", current_user.id, str(exc))
        return jsonify(
            success=False,
            message=f"Reset failed: {str(exc)}"
        ), 500


@admin_bp.route("/attendance/emergency-reset")
@login_required
@admin_required
def emergency_reset_attendance():
    """
    EMERGENCY: Delete all attendance via GET request (for immediate access).
    Can be accessed directly via browser URL.
    """
    import logging
    from app.models.attendance import Attendance
    from app.models.attendance_photo import AttendancePhoto
    from app.models.attendance_log import AttendanceLog
    from app.extensions.database import db

    logger = logging.getLogger("admin")

    try:
        # Count before deletion
        attendance_count = Attendance.query.count()
        photo_count = AttendancePhoto.query.count()
        log_count = AttendanceLog.query.count()

        logger.info(
            "EMERGENCY_ATTENDANCE_RESET | by_user=%s | att=%d | photos=%d | logs=%d",
            current_user.id, attendance_count, photo_count, log_count
        )

        # Delete in correct order
        AttendanceLog.query.delete()
        AttendancePhoto.query.delete()
        Attendance.query.delete()
        db.session.commit()

        logger.info(
            "EMERGENCY_RESET_SUCCESS | by_user=%s | deleted att=%d, photos=%d, logs=%d",
            current_user.id, attendance_count, photo_count, log_count
        )

        flash(
            f"✅ ATTENDANCE DELETED SUCCESSFULLY! "
            f"Deleted: {attendance_count} attendance records, {photo_count} photos, {log_count} logs.",
            "success"
        )
        return redirect(url_for("admin.index"))

    except Exception as exc:
        db.session.rollback()
        logger.error("EMERGENCY_RESET_FAILED | by_user=%s | error=%s", current_user.id, str(exc))
        flash(f"❌ ERROR: {str(exc)}", "danger")
        return redirect(url_for("admin.index"))


# ── Shift Assignment Routes ───────────────────────────────────────────

@admin_bp.route("/shift-assignment")
@login_required
@roles_required(UserRole.SUPER_ADMIN, UserRole.HR_MANAGER, UserRole.ADMIN)
def shift_assignment():
    """Bulk shift assignment page."""
    from .shift_assignment import assign_shifts_bulk
    return assign_shifts_bulk()


@admin_bp.route("/shift-assignment/assign", methods=["POST"])
@login_required
@roles_required(UserRole.SUPER_ADMIN, UserRole.HR_MANAGER, UserRole.ADMIN)
def assign_single_shift():
    """Assign shift to a single employee (AJAX)."""
    from .shift_assignment import assign_shift_to_employee
    return assign_shift_to_employee()


@admin_bp.route("/shift-assignment/bulk", methods=["POST"])
@login_required
@roles_required(UserRole.SUPER_ADMIN, UserRole.HR_MANAGER, UserRole.ADMIN)
def assign_bulk_shifts():
    """Bulk assign shifts (AJAX)."""
    from .shift_assignment import assign_shifts_bulk_submit
    return assign_shifts_bulk_submit()


@admin_bp.route("/shift-assignment/remove", methods=["POST"])
@login_required
@roles_required(UserRole.SUPER_ADMIN, UserRole.HR_MANAGER, UserRole.ADMIN)
def remove_shift():
    """Remove shift from employee (AJAX)."""
    from .shift_assignment import remove_shift_assignment
    return remove_shift_assignment()


@admin_bp.route("/shift-assignment/employee-info")
@login_required
@roles_required(UserRole.SUPER_ADMIN, UserRole.HR_MANAGER, UserRole.ADMIN)
def get_employee_shift():
    """Get employee shift info (AJAX)."""
    from .shift_assignment import get_employee_shift_info
    return get_employee_shift_info()


@admin_bp.route("/shift-assignment/assign-hospital", methods=["POST"])
@login_required
@roles_required(UserRole.SUPER_ADMIN, UserRole.HR_MANAGER, UserRole.ADMIN)
def assign_hospital():
    """Assign hospital to employee (AJAX)."""
    from .shift_assignment import assign_hospital_to_employee
    return assign_hospital_to_employee()


@admin_bp.route("/shifts/reseed", methods=["POST"])
@login_required
@roles_required(UserRole.SUPER_ADMIN)
def reseed_shifts():
    """Reseed all 25 shifts (SUPER_ADMIN only)."""
    from app.models.company import Shift  # noqa: PLC0415
    from app.extensions.database import db  # noqa: PLC0415
    
    try:
        # Clear existing shifts
        Shift.query.delete()
        db.session.commit()
        
        # Re-seed shifts
        from datetime import time as dt_time
        shifts_data = [
            {"name": "06:00 AM to 03:00 PM", "code": "SHIFT_0600_1500", "start_time": "06:00", "end_time": "15:00", "is_night": False},
            {"name": "06:30 AM to 03:30 PM", "code": "SHIFT_0630_1530", "start_time": "06:30", "end_time": "15:30", "is_night": False},
            {"name": "07:00 AM to 04:00 PM", "code": "SHIFT_0700_1600", "start_time": "07:00", "end_time": "16:00", "is_night": False},
            {"name": "07:30 AM to 04:30 PM", "code": "SHIFT_0730_1630", "start_time": "07:30", "end_time": "16:30", "is_night": False},
            {"name": "08:00 AM to 05:00 PM", "code": "SHIFT_0800_1700", "start_time": "08:00", "end_time": "17:00", "is_night": False},
            {"name": "08:00 AM to 06:00 PM", "code": "SHIFT_0800_1800", "start_time": "08:00", "end_time": "18:00", "is_night": False},
            {"name": "08:30 AM to 05:30 PM", "code": "SHIFT_0830_1730", "start_time": "08:30", "end_time": "17:30", "is_night": False},
            {"name": "09:00 AM to 06:00 PM", "code": "SHIFT_0900_1800", "start_time": "09:00", "end_time": "18:00", "is_night": False},
            {"name": "09:30 AM to 06:30 PM", "code": "SHIFT_0930_1830", "start_time": "09:30", "end_time": "18:30", "is_night": False},
            {"name": "10:00 AM to 06:00 PM", "code": "SHIFT_1000_1800", "start_time": "10:00", "end_time": "18:00", "is_night": False},
            {"name": "10:00 AM to 07:00 PM", "code": "SHIFT_1000_1900", "start_time": "10:00", "end_time": "19:00", "is_night": False},
            {"name": "10:15 AM to 07:15 PM", "code": "SHIFT_1015_1915", "start_time": "10:15", "end_time": "19:15", "is_night": False},
            {"name": "10:30 AM to 07:30 PM", "code": "SHIFT_1030_1930", "start_time": "10:30", "end_time": "19:30", "is_night": False},
            {"name": "11:00 AM to 08:00 PM", "code": "SHIFT_1100_2000", "start_time": "11:00", "end_time": "20:00", "is_night": False},
            {"name": "11:30 AM to 08:30 PM", "code": "SHIFT_1130_2030", "start_time": "11:30", "end_time": "20:30", "is_night": False},
            {"name": "12:00 PM to 09:00 PM", "code": "SHIFT_1200_2100", "start_time": "12:00", "end_time": "21:00", "is_night": False},
            {"name": "12:30 PM to 09:30 PM", "code": "SHIFT_1230_2130", "start_time": "12:30", "end_time": "21:30", "is_night": False},
            {"name": "12:45 PM to 09:45 PM", "code": "SHIFT_1245_2145", "start_time": "12:45", "end_time": "21:45", "is_night": False},
            {"name": "01:00 PM to 10:00 PM", "code": "SHIFT_1300_2200", "start_time": "13:00", "end_time": "22:00", "is_night": False},
            {"name": "01:00 PM to 06:00 PM", "code": "SHIFT_1300_1800", "start_time": "13:00", "end_time": "18:00", "is_night": False},
            {"name": "07:00 PM to 04:00 AM", "code": "SHIFT_1900_0400", "start_time": "19:00", "end_time": "04:00", "is_night": True},
            {"name": "09:00 PM to 06:00 AM", "code": "SHIFT_2100_0600", "start_time": "21:00", "end_time": "06:00", "is_night": True},
            {"name": "10:00 PM to 06:00 AM", "code": "SHIFT_2200_0600", "start_time": "22:00", "end_time": "06:00", "is_night": True},
            {"name": "10:00 PM to 07:00 AM", "code": "SHIFT_2200_0700", "start_time": "22:00", "end_time": "07:00", "is_night": True},
            {"name": "10:30 PM to 07:30 AM", "code": "SHIFT_2230_0730", "start_time": "22:30", "end_time": "07:30", "is_night": True},
        ]
        
        for shift_data in shifts_data:
            start_h, start_m = map(int, shift_data["start_time"].split(":"))
            end_h, end_m = map(int, shift_data["end_time"].split(":"))
            
            shift = Shift(
                name=shift_data["name"],
                code=shift_data["code"],
                start_time=dt_time(start_h, start_m),
                end_time=dt_time(end_h, end_m),
                is_night_shift=shift_data["is_night"],
                is_active=True,
                grace_minutes=10,
                break_minutes=60,
                working_days="Mon-Sun"
            )
            db.session.add(shift)
        
        db.session.commit()
        return jsonify({
            'success': True,
            'message': f'✅ Successfully seeded {len(shifts_data)} shifts',
            'count': len(shifts_data)
        })
    except Exception as e:
        db.session.rollback()
        import logging
        logger = logging.getLogger('admin')
        logger.error('Reseed shifts error: %s', str(e))
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500
