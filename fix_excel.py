#!/usr/bin/env python3
"""Fix Excel shift/hospital import file - normalize shift timings"""
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

# Define valid shifts
VALID_SHIFTS = [
    "06:00 AM to 03:00 PM",
    "06:30 AM to 03:30 PM",
    "07:00 AM to 04:00 PM",
    "07:30 AM to 04:30 PM",
    "08:00 AM to 05:00 PM",
    "08:00 AM to 06:00 PM",
    "08:30 AM to 05:30 PM",
    "09:00 AM to 06:00 PM",
    "09:30 AM to 06:30 PM",
    "10:00 AM to 06:00 PM",
    "10:00 AM to 07:00 PM",
    "10:15 AM to 07:15 PM",
    "10:30 AM to 07:30 PM",
    "11:00 AM to 08:00 PM",
    "11:30 AM to 08:30 PM",
    "12:00 PM to 09:00 PM",
    "12:30 PM to 09:30 PM",
    "12:45 PM to 09:45 PM",
    "01:00 PM to 10:00 PM",
    "01:00 PM to 06:00 PM",
    "07:00 PM to 04:00 AM",
    "09:00 PM to 06:00 AM",
    "10:00 PM to 06:00 AM",
    "10:00 PM to 07:00 AM",
    "10:30 PM to 07:30 AM",
]

def normalize_shift(shift_str):
    """Normalize shift string to match system format."""
    if not shift_str:
        return None
    
    shift_str = shift_str.strip()
    if not shift_str:
        return None
    
    # Replace periods with colons
    shift_str = shift_str.replace(".00", ":00").replace(".30", ":30").replace(".15", ":15").replace(".45", ":45")
    
    # Try to match time pattern: "HH:MM AM/PM to HH:MM AM/PM"
    pattern = r'(\d{1,2}):(\d{2})\s*(AM|PM|am|pm)\s+to\s+(\d{1,2}):(\d{2})\s*(AM|PM|am|pm)'
    match = re.match(pattern, shift_str)
    
    if match:
        start_h, start_m, start_p, end_h, end_m, end_p = match.groups()
        start_h = int(start_h)
        end_h = int(end_h)
        
        # Format as "HH:MM AM to HH:MM PM"
        normalized = f"{start_h:02d}:{start_m} {start_p.upper()} to {end_h:02d}:{end_m} {end_p.upper()}"
        
        # Check if it matches a valid shift (case-insensitive)
        for valid in VALID_SHIFTS:
            if normalized.lower() == valid.lower():
                return valid
        
        # No exact match found
        return None
    
    # Handle incomplete shift (just start time) - return None
    return None

# Load workbook
print("Reading Excel file...")
wb = load_workbook(r'c:\Users\durve\Downloads\bulk shift upload.xlsx')
ws = wb.active

# Create new workbook
new_wb = Workbook()
new_ws = new_wb.active
new_ws.title = "All Employee"

# Add headers
new_ws['B1'] = "EMP-CODE"
new_ws['C1'] = "HOSPITAL NAME"
new_ws['D1'] = "SHIFT"

# Style headers
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
for col in ['B', 'C', 'D']:
    new_ws[f'{col}1'].fill = header_fill
    new_ws[f'{col}1'].font = header_font

# Process rows
row_num = 2
fixed_count = 0
skipped_count = 0
error_count = 0

print("Processing rows...")
for src_row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
    # Get values
    emp_code_cell = row[1]  # Column B
    hospital_cell = row[2]  # Column C
    shift_cell = row[3]     # Column D
    
    emp_code = str(emp_code_cell.value).strip() if emp_code_cell.value else ""
    hospital = str(hospital_cell.value).strip() if hospital_cell.value else ""
    shift = str(shift_cell.value).strip() if shift_cell.value else ""
    
    # Skip empty rows
    if not emp_code or emp_code.lower() in ("emp-code", "none", "nan"):
        continue
    
    # Normalize shift
    shift_fixed = normalize_shift(shift)
    
    if not shift_fixed:
        print(f"  ✗ Row {src_row_num}: {emp_code} - Invalid shift: '{shift}'")
        skipped_count += 1
        continue
    
    # Write to new sheet
    new_ws[f'B{row_num}'] = emp_code
    new_ws[f'C{row_num}'] = hospital
    new_ws[f'D{row_num}'] = shift_fixed
    
    row_num += 1
    fixed_count += 1
    
    if fixed_count % 50 == 0:
        print(f"  ✓ Processed {fixed_count} rows...")

# Adjust column widths
new_ws.column_dimensions['B'].width = 15
new_ws.column_dimensions['C'].width = 40
new_ws.column_dimensions['D'].width = 25

# Save
output_file = r'c:\Users\durve\Downloads\bulk_shift_upload_FIXED.xlsx'
new_wb.save(output_file)

print(f"\n✅ DONE!")
print(f"   Fixed: {fixed_count} rows")
print(f"   Skipped: {skipped_count} rows (invalid shifts)")
print(f"   Output: {output_file}")
